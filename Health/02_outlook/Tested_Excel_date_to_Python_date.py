# %%
"""
現在のエクセルの日付はエクセル形式の日付となっているが、Pythonで罫線、日付その他の項目を入力しようとすると日付がうまく入らない。
従って、すべての日付をパイソンのdatetime形式のdateに変更しておくことにする
その場合の問題点
1)　一回だけ日付を変更するプログラムを起動する。日付の書式が混在しているケースをどうするか？
2）　日付を検索するプログラムはexcel日付を検索するようになっているが、datetimeの日付を検索するように変更する必要がある。
3)　日付の検索プログラムは複数ある
"""

# %%
# 20240830確認済　一度実行したので、2回する必要はない。20250615　食事記録表(Python月間190303).xlsxに実施
import openpyxl
import datetime
from datetime import datetime
from openpyxl import utils
from openpyxl import load_workbook
from openpyxl.styles import Alignment    # Alignmentクラスをインポート

from openpyxl import load_workbook
filepath="H:\Python\ExcelDATA\食事記録表(Python月間190303).xlsx"
wb=load_workbook(filename=filepath)
ws=wb["食事"]

# Excelファイルの最初からPythonの日付に書き換える
# 元の書式は　2018/12/23スタート

k=0
strr = 2 #スタートの行は2から
stc = 1 #スタートの列は1から
Date_to_be_changed = ws.cell(row=strr, column=stc).value
#print("start", Date_to_be_changed)

while Date_to_be_changed!=None: #空の日付のセルが検出されるまで
    k=k+1    
    if k>200:
        print("Stop")
        break
    
    col_index=ws.cell(row=strr, column=stc).column_letter
    ws.column_dimensions[col_index].width = "10" #幅を広げるhttps://note.com/freedom997/n/n36f17fa6f8e0?magazine_key=me477a581b514

    # 縦に11回繰り返す     
    for p in range(11):
        Date_to_be_changed = ws.cell(row=strr, column=stc).value
        Date_to_be_changed_1=utils.datetime.from_excel(Date_to_be_changed)  #Date_to_be_changedを年月日時刻に変形する
        #print(Date_to_be_changed_1)
        Date_to_be_changed_2=Date_to_be_changed_1.date()
        #print(Date_to_be_changed_2)
        ws.cell(row=strr, column=stc).number_format = 'yy/mm/dd'
        ws.cell(row=strr, column=stc).value = Date_to_be_changed_2
        date_ex = ws.cell(row=strr, column=stc).value
        #print("p:", p, "stc=", stc, "strr=", strr, "date=", Date_to_be_changed-2, date_ex)
        strr=strr+7

    stc=stc+7
    strr=2
    Date_to_be_changed = ws.cell(row=strr, column=stc).value
    #print("stc:", stc, type(stc))

else:
    print("Complete")

wb.save(filepath)

# %%
# 新しく罫線と項目名を追加する場合は、ここから起動する
# 空のセルを検出する2025/01/09　確認済

# %%
import pandas as pd
import openpyxl
from openpyxl import utils
from datetime import datetime
from openpyxl import load_workbook
import os

def get_excel_path(filename, folder="ExcelDATA"):
    base = os.path.join(os.environ["OneDrive"], "ドキュメント", "PythonWork")
    return os.path.join(base, folder, filename)

filepath = get_excel_path("食事記録表(--20240806-3).xlsx")

wb=load_workbook(filename=filepath)
ws=wb["食事"]

#食事記録表(Python用).xlsxの転記する日付を探し、そのエクセル上のアドレスを求める

table_Column_num=1 #線を引き始める日付の列をもとめる最初
k=0

Date_filled = ws.cell(row=2,column=1).value
print(Date_filled)
# 読み込んだ日付(Date_filled)は　datetime形式　2018-12-23 00:00:00 確認済　20240830

while Date_filled!=None:
    k=k+1    
    if k>400:
        print("Error")
        break
    else:
        table_Column_num=table_Column_num+7   #　線を引き始める日付の列
    Date_filled=ws.cell(row=2,column=table_Column_num-7).value
    
Date_filled=ws.cell(row=2,column=table_Column_num-14).value # 線が引いてある最後の先頭の日付
print("線が引いてある最後の先頭の日付", Date_filled)

print("線が引いてある最後の日付列番号: table_Column_num-7", table_Column_num-7)
Column_Address=utils.get_column_letter(table_Column_num-7)
print("Column_Address", Column_Address)

# %%
# 罫線と項目名を追加で入れる
# 2025/01/09確認済

# %%
# 2024/08/31　確認済　エクセルファイルに以前のデータの記憶が残っていた
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import datetime

# 黒い実線を引く
side1 = Side(style='thin', color='000000')

# 罫線を引く
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1) # セルの上下左右に罫線を引く
border_no = Border(top=side1, bottom=side1, left=None, right=None) # セルの左右の罫線がないもの
border_no_1 = Border(top=side1, bottom=side1, left=side1, right=None) # セルの右側の罫線のないもの
border_no_2 = Border(top=side1, bottom=side1, left=None, right=side1) # セルの左側の罫線のないもの
border_no_3 = Border(top=None, bottom=None, left=side1, right=side1) # セルの右と左のみ罫線を引く

strr = 2
stc = table_Column_num-7
print("1: strr:", strr, "stc:", stc)
# 線を引き始める最初の列は　table_Column_num-7
# 線を引き始める最初の行は　2

for m in range(6):   #縦11日分を一グループとし、それをｍグループ　線引き項目名を入れる
    strr=2  #線の引き始めは行2から始める


    for j in range(11):
        # 最初に日付のセルを結合
        ws.merge_cells(start_row=strr+j*7, start_column=stc,  end_row=strr+6+j*7, end_column=stc)
        col_index=ws.cell(row=strr+j*7, column=stc).column_letter
        ws.column_dimensions[col_index].width = "10" #幅を広げるhttps://note.com/freedom997/n/n36f17fa6f8e0?magazine_key=me477a581b514

    # 必要なセルに線を引く    
    for p in range(11): #11日分繰り返す
        for r in range(7): # 一日分が7行だから縦に7回繰り返す　行をひとつづつ
            for c in range(7): # 一日分は7列だから横に7セル分の線を引く　列を横にひとつづつ移動する
                if c==0 or c==1:
                    ws.cell(row=strr+r, column=stc+c).border = border_aro

                elif (c>1 and c<7) and (r==3 or r==4):
                    ws.cell(row=strr+r, column=stc+c).border = border_no_1
                    if c==6:
                        ws.cell(row=strr+r, column=stc+c).border = border_aro

                elif (c==3 or c==4)and (r==3 or r==4):
                    ws.cell(row=strr+r, column=stc+c).border = border_aro
                elif c==2 and r==5:
                    ws.cell(row=strr+r, column=stc+c).border = border_no_3
                elif (c==3 or c==4 or c==5) and r==5:
                    ws.cell(row=strr+r, column=stc+c).border = border_no_2
                elif c==6:
                    ws.cell(row=strr+r, column=stc+c).border = border_no_2

                else:
                    ws.cell(row=strr+r, column=stc+c).border = border_no 

                #print("r=", r, "strr+r", strr+r, "c=", c, "stc+c", stc+c)

        strr=strr+7 #行を7個下げる
                
    stc=stc+7 #列を7個右に移動する   

# セルに日付を挿入する
from openpyxl.styles import Alignment    # Alignmentクラスをインポート

strr = 2
stc = table_Column_num-7
print("2: strr:", strr, "stc:", stc)
# 線を引き始める最初の列は　table_Column_num
# 線を引き始める最初の行は　2

# date
# ここに書きこむ日付は、 Date_filled2+11days

delta1 = datetime.timedelta(days=1)
delta2 = datetime.timedelta(days=11)
Date3=Date_filled+delta2-delta1
Date_Data = datetime.date(Date3.year, Date3.month, Date3.day)
print(Date_Data)

fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='ADD8E6', bgColor='ADD8E6')

for m in range(6):  #縦11日分を一グループとし、それをｍグループ分繰り返す　
    strr=2  #文字挿入始めは行2から始める

    for p in range(11):  #ひとグループは11日
        for r in range(7): # 一日分は7行、縦に7回繰り返す
            for c in range(7): # 一日分は7列、横に7セル分移動する
                # 年月日を挿入
                if c==0 and r==0:
                    Date_Data = Date_Data+delta1
                    ws.cell(row=strr+r, column=stc+c).number_format = 'yy/mm/dd'
                    ws.cell(row=strr+r, column=stc+c).value = Date_Data
                    print("Date_Data=", Date_Data)

                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal='center') # 中央揃え
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(vertical='center')   # 全体表示(中央揃え)
                    if Date_Data.weekday() == 6: #青く塗りつぶす
                        ws.cell(row=strr+r, column=stc+c).fill = fill
                    
        strr=strr+7

    stc=stc+7


strr = 2
stc = table_Column_num-7
print("3: strr:", strr, "stc:", stc)
# 線を引き始める最初の列は　table_Column_num
# 線を引き始める最初の行は　2

# date
# ここに書きこむ日付は、 Date_filled2+11days
for m in range(6):   #6日分を一塊とし、それを3個　項目名を入れる
    strr=2  #文字挿入始めは行2から始める

    for p in range(11):
        for r in range(7): # 縦に7回繰り返す
            for c in range(7): # 横に7セル分移動する
                if r==0 and c==1:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="朝食"
                if r==1 and c==1:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="昼食"
                if r==2 and c==1:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="夕食"
                if r==3 and c==1:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="朝食前体重"
                if r==4 and c==1:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="薬服用"                    
                if r==5 and c==1:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="血糖値"
                if r==5 and c==3:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="歩数"
                if r==5 and c==5:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="消費Kcal"
                if r==6 and c==1:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="評価" 
                if r==3 and c==3:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="ウオーキング"                    
                if r==4 and c==3:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="早朝血圧"                    
                if r==3 and c==5:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="中程度運動"                    
                if r==4 and c==5:
                    ws.cell(row=strr+r, column=stc+c).alignment = Alignment(horizontal="centerContinuous")
                    ws.cell(row=strr+r, column=stc+c).value="飲酒"                                      
                    
        strr=strr+7
                
    stc=stc+7 #列を7個みぎに移動する 
                        
    
wb.save(filepath) 

# %%
# 上記のpython programuは　excel file H:\Python\ExcelDATA\食事記録表(--20230326)for_test4.xlsxで動作確認済み　24/08/31


