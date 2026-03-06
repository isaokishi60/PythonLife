# データの文字数によりエクセルの列の幅を調整する関数

import math
import datetime
import unicodedata
from openpyxl.utils import get_column_letter

def _visual_len(value):
    """Excel上の見た目の文字幅（英数=1、全角=2）を概算する"""
    if value is None:
        return 0
    if isinstance(value, (datetime.date, datetime.datetime)):
        # Dateは表示をyyyy-mm-ddに揃える（10文字）
        text = value.strftime("%Y-%m-%d")
    else:
        text = str(value)

    n = 0
    for ch in text:
        ea = unicodedata.east_asian_width(ch)
        # 全角/Wide(F,W)は2、その他は1として数える
        n += 2 if ea in ("F", "W") else 1
        # "+="" は、「加算して代入する」という意味の演算子
    return n

# 🔍 1文字ずつ処理
# text の中の文字 ch を1つずつ取り出して処理してる。
# 🔍 unicodedata.east_asian_width(ch)
# この関数は、文字の「幅の種類」を返す。
# 返ってくる値は以下のいずれか：

# "F"（Fullwidth）→ 全角
# "W"（Wide）→ 全角
# "Na"（Narrow）→ 半角
# "H"（Halfwidth）→ 半角
# "A"（Ambiguous）→ あいまい
# "N"（Neutral）→ 中立

# 🔍 幅のカウント
# "F" や "W" の場合は 2文字分の幅としてカウント。
# それ以外は 1文字分の幅としてカウント。

def autofit_columns(ws, header_row=1, padding=2, min_width=4):
    """ワークシートの列幅を日本語対応で自動調整"""
    # ヘッダ名 → 列番号 の辞書
    header_map = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row])}

    # Date列の表示形式を統一（あれば）
    if "Date" in header_map:
        c = header_map["Date"]
        for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row,
                                min_col=c, max_col=c):
            cell = row[0]
            cell.number_format = "yyyy-mm-dd"  # 表示は 2025-09-01（10文字）

    # 列ごとに最大見かけ幅を計算
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        is_ascii_only = True

        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is None:
                continue
            text = str(val)
            # 全角チェック
            for ch in text:
                ea = unicodedata.east_asian_width(ch)
                if ea in ("F", "W"):
                    is_ascii_only = False
            max_len = max(max_len, _visual_len(val))

        # 半角ばかりの列（IDや数字中心）は少し広めにする
        if is_ascii_only:
            width = max(min_width, math.ceil(max_len * 1.2 + padding))
        else:
            width = max(min_width, math.ceil(max_len + padding))

        ws.column_dimensions[get_column_letter(col_idx)].width = width
