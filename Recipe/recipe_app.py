import re
import os
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# =========================
# パス（OneDrive配下）
# =========================
BASE = Path(os.environ.get("OneDrive", "")) / "タブレット用"
DB_PATH = BASE / "recipe_db.xlsx"
SHOTS_ROOT = BASE / "RecipeShots"
SHEET_NAME = "Recipes"


# =========================
# ユーティリティ
# =========================
def norm(x) -> str:
    if x is None:
        return ""
    s = str(x).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def is_blank(x) -> bool:
    if x is None:
        return True
    try:
        if pd.isna(x):
            return True
    except Exception:
        pass
    return str(x).strip() == ""

def safe_str(x) -> str:
    if is_blank(x):
        return ""
    return str(x).strip()

def norm_rating(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return str(x).strip().upper()

def clean_url_for_mobile(u: str) -> str:
    u = (u or "").strip()
    if not u:
        return ""
    sp = urlsplit(u)
    # query（?以降）と fragment（#以降）を落とす
    return urlunsplit((sp.scheme, sp.netloc, sp.path, "", ""))

def list_images(folder: Path):
    exts = ["*.png", "*.jpg", "*.jpeg", "*.webp"]
    files = []
    for pat in exts:
        files.extend(folder.glob(pat))
    return sorted(files)

def split_genres(x) -> list[str]:
    """ジャンルセル '鍋, 肉' -> ['鍋','肉']（区切りゆれ吸収）"""
    if is_blank(x):
        return []
    s = str(x).strip().replace("、", ",")
    parts = [p.strip() for p in s.split(",") if p.strip()]
    return parts


# =========================
# Excel 書き戻し（ジャンル/評価）
# =========================
def save_genre_rating_to_excel(recipe_id: str, genres: list[str], rating: str) -> None:
    """
    Recipesシートの該当RecipeID行の
      ジャンル列（上書き）
      評価列（上書き：空もOK）
    を保存する
    """
    wb = load_workbook(DB_PATH)
    ws = wb[SHEET_NAME]

    header = [c.value for c in ws[1]]
    for col in ["RecipeID", "ジャンル", "評価"]:
        if col not in header:
            wb.close()
            raise ValueError(f"Excelに '{col}' 列が見つかりません（ヘッダー名を確認）")

    col_id = header.index("RecipeID") + 1
    col_genre = header.index("ジャンル") + 1
    col_rate = header.index("評価") + 1

    target_row = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_id).value
        if v is not None and str(v).strip() == str(recipe_id).strip():
            target_row = r
            break

    if target_row is None:
        wb.close()
        raise ValueError(f"RecipeID {recipe_id} が見つかりません")

    ws.cell(target_row, col_genre).value = ", ".join([g.strip() for g in genres if g.strip()])
    ws.cell(target_row, col_rate).value = (rating or "").strip().upper()

    wb.save(DB_PATH)
    wb.close()


# =========================
# データ読み込み
# =========================
@st.cache_data
def load_recipes():
    if not DB_PATH.exists():
        raise FileNotFoundError(f"DBが見つかりません: {DB_PATH}")

    df = pd.read_excel(DB_PATH, sheet_name=SHEET_NAME)

    # 必須列
    need = ["RecipeID", "レシピ名", "主材料1", "主材料2", "主材料3"]
    miss = [c for c in need if c not in df.columns]
    if miss:
        raise ValueError(f"{SHEET_NAME}に必要列がありません: {miss}")

    # 任意列（無くても動く）
    for c in ["評価", "ジャンル", "URL"]:
        if c not in df.columns:
            df[c] = ""

    # RecipeID を文字列化
    df["RecipeID"] = df["RecipeID"].astype(str).str.strip()

    return df


# =========================
# フィルター
# =========================
def filter_recipes(
    df: pd.DataFrame,
    kw: str,
    genre_sel: list[str],
    rating_opt: str,
    mat1: str,
    mat2: str,
    mat_mode: str,
) -> pd.DataFrame:

    # ここが必要（フィルタの土台）
    f = pd.Series(True, index=df.index)

    # 評価（現行仕様を維持）
    rr = df["評価"].map(norm_rating)
    if rating_opt == "選択しない":
        pass
    elif rating_opt == "全て":
        f &= rr.isin(["A", "B", "C"])
    else:
        f &= rr.eq(rating_opt)

    # ジャンル（複数選択）
    if genre_sel:
        g_lists = df["ジャンル"].map(split_genres)
        f &= g_lists.apply(lambda gs: any(g in gs for g in genre_sel))

    # レシピ名（部分一致）
    if kw:
        f &= df["レシピ名"].astype(str).map(norm).str.contains(re.escape(norm(kw)), na=False)

    # 主材料（2入力＋AND/OR）
    mat1 = (mat1 or "").strip()
    mat2 = (mat2 or "").strip()
    mat_mode = (mat_mode or "AND").strip().upper()

    def has_material(term: str) -> pd.Series:
        v = re.escape(norm(term))
        return (
            df["主材料1"].astype(str).map(norm).str.contains(v, na=False)
            | df["主材料2"].astype(str).map(norm).str.contains(v, na=False)
            | df["主材料3"].astype(str).map(norm).str.contains(v, na=False)
        )

    if mat1 or mat2:
        if mat1 and mat2:
            if mat_mode == "AND":
                f &= has_material(mat1) & has_material(mat2)
            else:  # OR
                f &= has_material(mat1) | has_material(mat2)
        else:
            f &= has_material(mat1 or mat2)

    return df[f].copy()




# =========================
# ページ状態
# =========================
if "page" not in st.session_state:
    st.session_state.page = "search"   # search / select / view / tag

# 検索条件の保持
st.session_state.setdefault("kw", "")
st.session_state.setdefault("m", "")
st.session_state.setdefault("genre_sel", [])
st.session_state.setdefault("rating_opt", "選択しない")
st.session_state.setdefault("selected_ids", [])
st.session_state.setdefault("tag_rid", "")


# =========================
# UI
# =========================
st.set_page_config(page_title="レシピ検索", layout="wide")
st.title("レシピ検索（材料表＋調理法）")

df = load_recipes().copy()

# ジャンル候補をDBから自動生成
ALL_GENRES = sorted({g for x in df["ジャンル"] for g in split_genres(x)})

# 共通：検索結果
def get_res():
    return filter_recipes(
        df,
        st.session_state.kw,
        st.session_state.genre_sel,
        st.session_state.rating_opt,
        st.session_state.get("mat1", ""),
        st.session_state.get("mat2", ""),
        st.session_state.get("mat_mode", "AND"),
    )



# =========================
# ページ：search
# =========================
if st.session_state.page == "search":
    st.subheader("① 検索条件")

    kw = st.text_input("レシピ名（部分一致）", value=st.session_state.kw)
    st.subheader("主材料検索")

    col1, col2, col3 = st.columns([3, 3, 2])

    with col1:
        mat1 = st.text_input(
            "主材料①",
            value=st.session_state.get("mat1", "")
        )

    with col2:
        mat2 = st.text_input(
            "主材料②（任意）",
            value=st.session_state.get("mat2", "")
        )

    with col3:
        mat_mode = st.radio(
            "条件",
            ["AND", "OR"],
            horizontal=True,
            index=0 if st.session_state.get("mat_mode", "AND") == "AND" else 1,
        )

    # セッション保存（ページ切替や再描画対策）
    st.session_state.mat1 = mat1
    st.session_state.mat2 = mat2
    st.session_state.mat_mode = mat_mode


    genre_sel = st.multiselect(
        "ジャンル（複数選択可）",
        options=ALL_GENRES,
        default=st.session_state.genre_sel,
    )

    rating_opt = st.selectbox(
        "評価（作ったレシピ）",
        ["選択しない", "全て", "A", "B", "C"],
        index=["選択しない", "全て", "A", "B", "C"].index(st.session_state.rating_opt),
    )

    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button("検索する", use_container_width=True):
            st.session_state.kw = kw
            st.session_state.genre_sel = genre_sel
            st.session_state.rating_opt = rating_opt

            # ---- 主材料（新方式） ----
            st.session_state.mat1 = mat1
            st.session_state.mat2 = mat2
            st.session_state.mat_mode = mat_mode

            st.session_state.page = "select"
            st.session_state.selected_ids = []

            st.rerun()


    with c2:
        if st.button("条件クリア", use_container_width=True):
            st.session_state.kw = ""

            # ---- 主材料（新方式） ----
            st.session_state.mat1 = ""
            st.session_state.mat2 = ""
            st.session_state.mat_mode = "AND"

            st.session_state.genre_sel = []
            st.session_state.rating_opt = "選択しない"
            st.session_state.selected_ids = []

            st.rerun()


    with c3:
        if st.button("ジャンル登録（R番号）", use_container_width=True):
            st.session_state.page = "tag"
            st.rerun()

    st.caption("※ まず「ジャンル」で絞ってから、名前/材料で探すと500件でも速いです。")
    st.stop()

# =========================
# ページ：select
# =========================
if st.session_state.page == "select":
    st.subheader("② 候補から選択")

    res = get_res()
    st.caption(f"該当: {len(res)}件")

    if len(res) == 0:
        st.info("条件に一致するレシピがありません。")
        if st.button("← 検索へ戻る"):
            st.session_state.page = "search"
            st.rerun()
        st.stop()

    show_cols = ["RecipeID", "レシピ名", "主材料1", "主材料2", "主材料3", "評価", "ジャンル"]
    st.dataframe(res[show_cols], hide_index=True, use_container_width=True)

    # 複数選択（タブレット安定）
    def make_label(r):
        rid = safe_str(r.get("RecipeID", ""))
        rate = norm_rating(r.get("評価", ""))
        name = safe_str(r.get("レシピ名", ""))
        rate_disp = rate if rate else "-"
        return f"{rid} | {rate_disp} | {name}"

    options = res["RecipeID"].astype(str).tolist()
    label_map = {rid: make_label(res[res["RecipeID"] == rid].iloc[0]) for rid in options}

    selected = st.multiselect(
        "表示するレシピ（複数選択可）",
        options=options,
        default=st.session_state.selected_ids,
        format_func=lambda x: label_map.get(x, x),
    )
    st.session_state.selected_ids = selected

    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("← 検索へ戻る", use_container_width=True):
            st.session_state.page = "search"
            st.rerun()
    with c2:
        if st.button("ジャンル登録（R番号）", use_container_width=True):
            st.session_state.page = "tag"
            st.rerun()
    with c3:
        if st.button("表示する", use_container_width=True):
            if not selected:
                st.warning("表示するレシピを選んでください。")
            else:
                st.session_state.page = "view"
                st.rerun()

    st.stop()

# =========================
# ページ：view
# =========================
if st.session_state.page == "view":
    st.subheader("③ レシピ表示")

    res = get_res()
    selected = st.session_state.selected_ids

    if not selected:
        st.info("表示するレシピが未選択です。")
        if st.button("← 候補へ戻る"):
            st.session_state.page = "select"
            st.rerun()
        st.stop()

    c1, c2 = st.columns(2)
    with c1:
        if st.button("← 候補へ戻る", use_container_width=True):
            st.session_state.page = "select"
            st.rerun()
    with c2:
        if st.button("ジャンル登録（R番号）", use_container_width=True):
            st.session_state.page = "tag"
            st.rerun()

    for rid in selected:
        hit = res[res["RecipeID"].astype(str) == str(rid)]
        if hit.empty:
            continue
        row = hit.iloc[0]

        st.divider()
        st.subheader(f"{safe_str(row['レシピ名'])}（{rid}）")
        st.write(
            f"主材料: {safe_str(row.get('主材料1',''))} / "
            f"{safe_str(row.get('主材料2',''))} / "
            f"{safe_str(row.get('主材料3',''))}"
        )
        st.write(f"評価: {norm_rating(row.get('評価','')) or '（未入力）'}")
        st.write(f"ジャンル: {', '.join(split_genres(row.get('ジャンル',''))) or '（未入力）'}")

        # URL（参考）
        url_raw = safe_str(row.get("URL", ""))
        url = clean_url_for_mobile(url_raw)
        if url:
            if "kurashiru.com" in url:
                st.caption("※ クラシルはタブレットでは会員/誘導表示になる場合があります（画像が正本です）。")
            st.link_button("参考URLを開く", url)
        else:
            st.write("URL: URLはない")

        # 画像
        rdir = SHOTS_ROOT / str(rid)
        if not rdir.exists():
            st.error(f"画像フォルダが見つかりません: {rdir}")
            continue

        imgs = list_images(rdir)
        if not imgs:
            st.warning(f"{rdir} に画像がありません。")
            continue

        st.info("上から順に表示します。最初を材料、続きが調理法になるように並び（ファイル名）を揃えると便利です。")
        for p in imgs:
            st.image(str(p), caption=p.name, use_container_width=True)

    st.stop()


# =========================
# ページ：tag（R番号でジャンル/評価を編集）
# =========================
if st.session_state.page == "tag":
    st.subheader("ジャンル登録（R番号で編集）")

    rid = st.text_input("RecipeID（例：R123）", value=st.session_state.get("tag_rid", "")).strip()
    st.session_state.tag_rid = rid

    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("← 検索へ戻る", use_container_width=True):
            st.session_state.page = "search"
            st.rerun()
    with c2:
        if st.button("← 候補へ戻る", use_container_width=True):
            st.session_state.page = "select"
            st.rerun()
    with c3:
        if st.button("← 表示へ戻る", use_container_width=True):
            st.session_state.page = "view"
            st.rerun()

    if not rid:
        st.info("RecipeIDを入力してください。")
        st.stop()

    hit = df[df["RecipeID"].astype(str) == rid]
    if hit.empty:
        st.error(f"{rid} が見つかりません。")
        st.stop()

    row = hit.iloc[0]
    st.subheader(safe_str(row["レシピ名"]))

    # 現在値
    current_genres = split_genres(row.get("ジャンル", ""))
    current_rating = norm_rating(row.get("評価", ""))

    # ジャンル選択
    sel = st.multiselect("ジャンル（複数選択可）", options=ALL_GENRES, default=current_genres)

    # 新ジャンル追加（任意）
    newg = st.text_input("新しいジャンルを追加（任意）", "")
    if newg.strip():
        ng = newg.strip()
        if ng not in sel:
            sel = sel + [ng]
        # その場で候補にも反映（ただし保存はDB側）
        if ng not in ALL_GENRES:
            ALL_GENRES.append(ng)
            ALL_GENRES.sort()

    # 評価（現行どおり：A/B/C/空）
    rating = st.selectbox(
        "評価（A/B/C または空）",
        options=["", "A", "B", "C"],
        index=["", "A", "B", "C"].index(current_rating) if current_rating in ["A","B","C"] else 0,
    )

    st.caption("※ Excel( recipe_db.xlsx )を開いていると保存に失敗することがあります。保存時は閉じてください。")

    if st.button("保存（Excelへ反映）", use_container_width=True):
        try:
            save_genre_rating_to_excel(rid, sel, rating)
            st.success("保存しました。")

            # ---- 追加：入力R番号をクリア ----
            st.session_state.tag_rid = ""

            # DB再読み込み
            st.cache_data.clear()

            # 画面をリフレッシュ
            st.rerun()

        except Exception as e:
            st.error(f"保存に失敗しました: {e}")


    st.stop()
