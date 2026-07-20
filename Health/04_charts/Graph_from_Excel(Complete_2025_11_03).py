# データ表1.xlsx の Sheet4 から、指定項目の健康グラフを1つ作成する
# Power Automate対応版
#
# 実行例:
# python "Graph_from_Excel(Complete_2025_11_03).py" --start-date 2025-10-01 --end-date 2026-06-17 --item 1
import os
import argparse
import datetime

from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart
from openpyxl.chart.axis import DateAxis
from openpyxl.utils import get_column_letter

import pandas as pd

# =========================
# 引数
# =========================

parser = argparse.ArgumentParser()
parser.add_argument("--start-date", required=True, help="開始日 YYYY-MM-DD")
parser.add_argument("--end-date", required=True, help="終了日 YYYY-MM-DD")
parser.add_argument("--item", required=True, type=int, help="1:体重 2:血糖値 3:血圧 4:中程度運動 5:運動消費 6:歩数")

args = parser.parse_args()

real_start = datetime.datetime.strptime(args.start_date, "%Y-%m-%d").date()
real_end = datetime.datetime.strptime(args.end_date, "%Y-%m-%d").date()
Item_input = args.item

MOVING_AVERAGE_DAYS = 7

moving_average_data_start = (
    real_start
    - datetime.timedelta(days=MOVING_AVERAGE_DAYS - 1)
)

DayOne = datetime.timedelta(days=1)

# 元コードと同じく、不等号判定用に前後1日ずらす
dt_start_input = real_start - DayOne
dt_end_input = real_end + DayOne

print("開始日", real_start)
print("終了日", real_end)
print("項目", Item_input)

if real_end < real_start:
    raise ValueError("終了日が開始日より前です")

if Item_input not in [1, 2, 3, 4, 5, 6]:
    raise ValueError("item は 1～6 を指定してください")


# =========================
# パス設定
# =========================

def get_excel_path(filename, folder="ExcelDATA"):
    base = os.path.join(os.environ["OneDrive"], "ドキュメント", "PythonWork")
    return os.path.join(base, folder, filename)


filepath = get_excel_path("データ表1.xlsx", folder="ExcelDATA")


# =========================
# 項目定義
# =========================

ITEMS = {
    1: {
        "name": "体重",
        "headers": ["体重"],
        "sheet4_cols": [2],
        "ylabel": "体重",
    },
    2: {
        "name": "血糖値",
        "headers": ["血糖値"],
        "sheet4_cols": [3],
        "ylabel": "血糖値",
    },
    3: {
        "name": "血圧",
        "headers": ["血圧収縮期", "血圧拡張期", "心拍数"],
        "sheet4_cols": [4, 5, 6],
        "ylabel": "血圧",
    },
    4: {
        "name": "中程度運動",
        "headers": ["中程度運動"],
        "sheet4_cols": [7],
        "ylabel": "中程度運動(分)",
    },
    5: {
        "name": "運動消費エネルギー",
        "headers": ["運動消費エネルギー"],
        "sheet4_cols": [8],
        "ylabel": "運動消費カロリー",
    },
    6: {
        "name": "歩数",
        "headers": ["歩数"],
        "sheet4_cols": [9],
        "ylabel": "歩数",
    },
}

ITEM = ITEMS[Item_input]


# =========================
# Excel読み込み
# =========================

wb = load_workbook(filename=filepath)

if "Sheet4" not in wb.sheetnames:
    raise ValueError("データ表1.xlsx に Sheet4 がありません")

ws4 = wb["Sheet4"]

# Sheet2 / Sheet3 を毎回作り直す
for sheet_name in ["Sheet2", "Sheet3"]:
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

ws2 = wb.create_sheet("Sheet2")
ws3 = wb.create_sheet("Sheet3")

ws2.column_dimensions["A"].width = 12
ws3.column_dimensions["A"].width = 12


# =========================
# Sheet2 見出し
# =========================

ws2.cell(row=1, column=1).value = "日付"

for i, header in enumerate(ITEM["headers"], start=2):
    ws2.cell(row=1, column=i).value = header
    ws2.column_dimensions[get_column_letter(i)].width = 15


# =========================
# Sheet4 → Sheet2 転記
# =========================

read_row = 2
write_row = 2
k = 0

while True:
    k += 1
    if k > 5000:
        print("5000行を超えたため停止")
        break

    date_value = ws4.cell(row=read_row, column=1).value

    if date_value is None:
        break

    if isinstance(date_value, datetime.datetime):
        date3 = date_value.date()
    elif isinstance(date_value, datetime.date):
        date3 = date_value
    else:
        read_row += 1
        continue

    if moving_average_data_start <= date3 <= real_end:
        ws2.cell(row=write_row, column=1).value = date3

        for i, src_col in enumerate(ITEM["sheet4_cols"], start=2):
            ws2.cell(row=write_row, column=i).value = ws4.cell(row=read_row, column=src_col).value

        write_row += 1

    read_row += 1

print("転記行数:", write_row - 2)

# ============================================================
# 7日移動平均を作成
# ============================================================

def excel_value_to_float(value):
    """
    Excelセルの値を数値へ変換する。
    空欄や数値化できない値はNoneとする。
    0は有効な測定値として残す。
    """
    if value is None or value == "":
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def calculate_calendar_moving_average(
    dates,
    values,
    window_days=7,
):
    """
    暦日基準の移動平均を計算する。

    各日について、その日を含む直前window_days日間を対象とする。
    欠測値は平均から除外するが、0は平均に含める。
    """
    series = pd.Series(
        values,
        index=pd.to_datetime(dates),
        dtype="float64",
    )

    # 日付順に並べ、同じ日付が重複した場合は最後の値を使う
    series = series.sort_index()
    series = series[
        ~series.index.duplicated(keep="last")
    ]

    moving_average = series.rolling(
        f"{window_days}D",
        min_periods=1,
    ).mean()

    result = []

    for current_date in pd.to_datetime(dates):
        value = moving_average.get(current_date)

        if pd.isna(value):
            result.append(None)
        else:
            result.append(float(value))

    return result


data_row_start = 2
data_row_end = ws2.max_row

if data_row_end < data_row_start:
    raise RuntimeError(
        "指定期間のグラフ用データがありません。"
    )

moving_average_source_columns = []

if int(Item_input) in (1, 2, 4, 5, 6):
    # 単独項目：B列
    moving_average_source_columns = [2]

elif int(Item_input) == 3:
    # 血圧・心拍数：B～D列
    moving_average_source_columns = [2, 3, 4]


dates_for_average = [
    ws2.cell(row=row, column=1).value
    for row in range(data_row_start, data_row_end + 1)
]

# 元データの右側に7日移動平均列を追加する
moving_average_start_column = (
    max(moving_average_source_columns) + 1
)

for output_offset, source_column in enumerate(
    moving_average_source_columns
):
    output_column = (
        moving_average_start_column + output_offset
    )

    source_title = ws2.cell(
        row=1,
        column=source_column,
    ).value

    ws2.cell(
        row=1,
        column=output_column,
    ).value = f"{source_title} 7日移動平均"

    source_values = [
        excel_value_to_float(
            ws2.cell(
                row=row,
                column=source_column,
            ).value
        )
        for row in range(
            data_row_start,
            data_row_end + 1,
        )
    ]

    moving_values = calculate_calendar_moving_average(
        dates_for_average,
        source_values,
        MOVING_AVERAGE_DAYS,
    )

    for row, moving_value in zip(
        range(data_row_start, data_row_end + 1),
        moving_values,
    ):
        ws2.cell(
            row=row,
            column=output_column,
        ).value = moving_value

print(
    f"{MOVING_AVERAGE_DAYS}日移動平均作成完了: "
    f"{len(dates_for_average)}日分"
)

if write_row <= 2:
    raise ValueError("指定期間のデータがありません")


# ============================================================
# Sheet2からSheet3へグラフ用データを転記
# ============================================================

if int(Item_input) in (1, 2, 4, 5, 6):
    # 日付、日次値、7日移動平均
    graph_column_count = 3

elif int(Item_input) == 3:
    # 日付、収縮期、拡張期、心拍数、
    # 各項目の7日移動平均
    graph_column_count = 7

else:
    raise ValueError(
        f"項目番号が不正です: {Item_input}"
    )


# 見出しを転記
for column in range(1, graph_column_count + 1):
    ws3.cell(row=1, column=column).value = ws2.cell(
        row=1,
        column=column,
    ).value

# 指定期間だけをSheet3へ転記
sheet3_row = 2

for source_row in range(2, ws2.max_row + 1):
    source_date = ws2.cell(
        row=source_row,
        column=1,
    ).value

    if isinstance(source_date, datetime.datetime):
        source_date = source_date.date()

    if not isinstance(source_date, datetime.date):
        continue

    if real_start <= source_date <= real_end:
        for column in range(1, graph_column_count + 1):
            ws3.cell(
                row=sheet3_row,
                column=column,
            ).value = ws2.cell(
                row=source_row,
                column=column,
            ).value

        sheet3_row += 1


ws3.column_dimensions["A"].width = 12

for row in range(2, ws3.max_row + 1):
    ws3.cell(
        row=row,
        column=1,
    ).number_format = "yyyy-mm-dd"


# =========================
# 折れ線グラフ作成
# =========================

# ============================================================
# 折れ線グラフを作成
# ============================================================

graph_obj = LineChart()

graph_obj.style = 13
graph_obj.height = 16
graph_obj.width = 30
graph_obj.anchor = "F8"

graph_obj.title = (
    f"{ITEM['name']}"
    f"({real_start}～{real_end})"
)

graph_obj.x_axis = DateAxis(crossAx=500)
graph_obj.x_axis.number_format = "yyyy-mm-dd"
graph_obj.x_axis.title = "年月日"
graph_obj.x_axis.majorTimeUnit = "days"

graph_obj.y_axis.crossAx = 500

if int(Item_input) == 1:
    graph_obj.y_axis.title = "体重（kg）"

elif int(Item_input) == 2:
    graph_obj.y_axis.title = "血糖値（mg/dL）"

elif int(Item_input) == 3:
    graph_obj.y_axis.title = "血圧・心拍数"

elif int(Item_input) == 4:
    graph_obj.y_axis.title = "中程度運動量（分）"

elif int(Item_input) == 5:
    graph_obj.y_axis.title = (
        "運動消費エネルギー（kcal）"
    )

elif int(Item_input) == 6:
    graph_obj.y_axis.title = "歩数"


categories = Reference(
    ws3,
    min_col=1,
    min_row=2,
    max_row=ws3.max_row,
)


if int(Item_input) in (1, 2, 4, 5, 6):
    # B列：日次値
    # C列：7日移動平均
    daily_data = Reference(
        ws3,
        min_col=2,
        max_col=2,
        min_row=1,
        max_row=ws3.max_row,
    )

    average_data = Reference(
        ws3,
        min_col=3,
        max_col=3,
        min_row=1,
        max_row=ws3.max_row,
    )

    graph_obj.add_data(
        daily_data,
        titles_from_data=True,
    )

    graph_obj.add_data(
        average_data,
        titles_from_data=True,
    )

else:
    # 血圧・心拍数
    # B～D列：日次値
    # E～G列：7日移動平均
    daily_data = Reference(
        ws3,
        min_col=2,
        max_col=4,
        min_row=1,
        max_row=ws3.max_row,
    )

    average_data = Reference(
        ws3,
        min_col=5,
        max_col=7,
        min_row=1,
        max_row=ws3.max_row,
    )

    graph_obj.add_data(
        daily_data,
        titles_from_data=True,
        from_rows=False,
    )

    graph_obj.add_data(
        average_data,
        titles_from_data=True,
        from_rows=False,
    )


graph_obj.set_categories(categories)
graph_obj.legend.position = "b"


# ============================================================
# 系列の表示形式
# ============================================================

# 日次値と移動平均の対応色
SERIES_COLORS = [
    "C00000",  # 赤
    "4472C4",  # 青
    "70AD47",  # 緑
]

if int(Item_input) in (1, 2, 4, 5, 6):
    daily_series_count = 1
else:
    daily_series_count = 3


for index, series in enumerate(graph_obj.series):
    is_moving_average = (
        index >= daily_series_count
    )

    metric_index = (
        index - daily_series_count
        if is_moving_average
        else index
    )

    color = SERIES_COLORS[
        metric_index % len(SERIES_COLORS)
    ]

    series.graphicalProperties.line.solidFill = (
        color
    )

    # Excelのスムージングは使わない
    series.smooth = False

    if is_moving_average:
        # 7日移動平均：太線、マーカーなし
        series.graphicalProperties.line.width = 35000
        series.marker.symbol = "none"

    else:
        # 日次値：細線、小さい丸マーカー
        series.graphicalProperties.line.width = 10000
        series.marker.symbol = "circle"
        series.marker.size = 3

        series.marker.graphicalProperties.line.solidFill = (
            color
        )
        series.marker.graphicalProperties.solidFill = (
            color
        )


ws3.add_chart(graph_obj)


# =========================
# 保存
# =========================

wb.save(filepath)
wb.close()

print("グラフ作成完了:", ITEM["name"])
print(filepath)



