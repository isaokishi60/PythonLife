# -*- coding: utf-8 -*-
import os
from datetime import datetime, timedelta, date
from pathlib import Path

import pandas as pd
from garminconnect import Garmin

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

import matplotlib.pyplot as plt


# =========================
# 設定
# =========================
BASE_DIR = Path(__file__).resolve().parent          # ...\01_Garmin_Import
OUT_DIR  = BASE_DIR / "outputs"
OUT_DIR.mkdir(parents=True, exist_ok=True)

OUT_XLSX = OUT_DIR / "心拍数データ（警報付き）.xlsx"


SHEET_RAW = "心拍数データ"
SHEET_MONTH = "月別集計"
SHEET_CHART = "グラフ"

DEFAULT_START = date(2025, 10, 1)
DEFAULT_END = date.today()


def input_date(prompt: str, default: date) -> date:
    s = input(f"{prompt} [Enterで{default.isoformat()}]: ").strip()
    if not s:
        return default
    return datetime.strptime(s, "%Y-%m-%d").date()


def safe_int(x) -> int:
    try:
        return int(x)
    except Exception:
        return 0


def build_alert_level(rhr: int, rhr7: int, hr_max: int, hr_min: int) -> int:
    """
    0=通常, 1=注意(黄), 2=強い注意(赤)
    ※医療診断ではなく、セルフ管理用の注意信号
    """
    Y_RHR_DELTA = 8
    R_RHR_DELTA = 12
    Y_MAX = 120
    R_MAX = 135
    Y_MIN = 38
    R_MIN = 35

    level = 0

    if rhr7 > 0:
        delta = rhr - rhr7
        if delta >= R_RHR_DELTA:
            level = max(level, 2)
        elif delta >= Y_RHR_DELTA:
            level = max(level, 1)

    if hr_max >= R_MAX:
        level = max(level, 2)
    elif hr_max >= Y_MAX:
        level = max(level, 1)

    if hr_min > 0:
        if hr_min <= R_MIN:
            level = max(level, 2)
        elif hr_min <= Y_MIN:
            level = max(level, 1)

    return level


def fetch_garmin_daily_rows(start_d: date, end_d: date) -> pd.DataFrame:
    email = os.environ["GARMIN_EMAIL"]
    password = os.environ["GARMIN_PASSWORD"]

    g = Garmin(email, password)
    g.login()

    rows = []
    prev_rhr = None

    cur = start_d
    while cur <= end_d:
        ds = cur.isoformat()

        try:
            hr = g.get_heart_rates(ds) or {}
        except Exception as e:
            print(f"⚠️ {ds} 心拍取得に失敗: {e}")
            hr = {}

        rhr = safe_int(hr.get("restingHeartRate") or 0)
        hr_max = safe_int(hr.get("maxHeartRate") or 0)
        hr_min = safe_int(hr.get("minHeartRate") or 0)
        rhr7 = safe_int(hr.get("lastSevenDaysAvgRestingHeartRate") or 0)

        # 前日差
        if prev_rhr is None or rhr == 0:
            diff_prev = 0
        else:
            diff_prev = rhr - prev_rhr
        if rhr != 0:
            prev_rhr = rhr

        # 0はデータ無しとして空欄に
        max_out = None if hr_max == 0 else hr_max
        min_out = None if hr_min == 0 else hr_min

        level = build_alert_level(rhr=rhr, rhr7=rhr7, hr_max=hr_max, hr_min=hr_min)

        rows.append({
            "日付": cur,
            "安静時心拍数": rhr if rhr != 0 else None,
            "前日差": diff_prev if rhr != 0 else None,
            "最大心拍数": max_out,
            "最小心拍数": min_out,
            "RHR７日移動平均": rhr7 if rhr7 != 0 else None,
            "警報レベル": level,  # 0/1/2
        })

        print(f"✅ {ds} 取得: RHR={rhr} MAX={hr_max} MIN={hr_min} 7d={rhr7} level={level}")
        cur += timedelta(days=1)

    df = pd.DataFrame(rows)
    df["日付"] = pd.to_datetime(df["日付"])
    return df


def upsert_raw_to_excel(df_new: pd.DataFrame) -> pd.DataFrame:
    """既存があれば読み込み、日付で上書き統合（重複なし）"""
    if OUT_XLSX.exists():
        try:
            df_old = pd.read_excel(OUT_XLSX, sheet_name=SHEET_RAW)
            df_old["日付"] = pd.to_datetime(df_old["日付"])
        except Exception:
            df_old = pd.DataFrame(columns=df_new.columns)
    else:
        df_old = pd.DataFrame(columns=df_new.columns)

    df_all = pd.concat([df_old, df_new], ignore_index=True)
    df_all = df_all.sort_values("日付")
    df_all = df_all.drop_duplicates(subset=["日付"], keep="last").reset_index(drop=True)
    return df_all


def make_monthly_summary(df_all: pd.DataFrame) -> pd.DataFrame:
    work = df_all.copy()

    # ★数値列を強制的に数値化（文字列や空欄が混じってもOKにする）
    num_cols = ["安静時心拍数", "前日差", "最大心拍数", "最小心拍数", "RHR７日移動平均", "警報レベル"]
    for c in num_cols:
        if c in work.columns:
            work[c] = pd.to_numeric(work[c], errors="coerce")

    work["月"] = pd.to_datetime(work["日付"]).dt.to_period("M").astype(str)

    agg = work.groupby("月", as_index=False).agg(
        平均RHR=("安静時心拍数", "mean"),
        平均7d=("RHR７日移動平均", "mean"),
        月最大MAX=("最大心拍数", "max"),
        月最小MIN=("最小心拍数", "min"),
        赤日数=("警報レベル", lambda s: int((s == 2).sum())),
        黄日数=("警報レベル", lambda s: int((s == 1).sum())),
        通常日数=("警報レベル", lambda s: int((s == 0).sum())),
        日数=("日付", "count"),
    )

    # ★round前に念のため数値化（これでTypeErrorを確実に潰す）
    for col in ["平均RHR", "平均7d"]:
        agg[col] = pd.to_numeric(agg[col], errors="coerce").round(1)

    return agg


def write_excel(df_all: pd.DataFrame, df_month: pd.DataFrame) -> None:

    if OUT_XLSX.exists():
        # 既存ファイル → 追記モード（シート置換OK）
        with pd.ExcelWriter(
            OUT_XLSX,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as w:
            df_all.to_excel(w, sheet_name=SHEET_RAW, index=False)
            df_month.to_excel(w, sheet_name=SHEET_MONTH, index=False)

    else:
        # 新規ファイル → 書き込みモード（if_sheet_exists不要）
        with pd.ExcelWriter(
            OUT_XLSX,
            engine="openpyxl",
            mode="w"
        ) as w:
            df_all.to_excel(w, sheet_name=SHEET_RAW, index=False)
            df_month.to_excel(w, sheet_name=SHEET_MONTH, index=False)

    make_and_insert_charts(df_all)


def make_and_insert_charts(df_all: pd.DataFrame) -> None:
    df = df_all.copy().sort_values("日付")
    fig1_path = OUT_XLSX.with_suffix(".rhr.png")

    plt.figure()
    plt.plot(df["日付"], df["安静時心拍数"], label="RHR")
    plt.plot(df["日付"], df["RHR７日移動平均"], label="RHR 7d")
    plt.legend()
    plt.title("Resting Heart Rate (RHR) and 7-day Avg")
    plt.xlabel("Date")
    plt.ylabel("bpm")
    plt.tight_layout()
    plt.savefig(fig1_path, dpi=160)
    plt.close()

    wb = load_workbook(OUT_XLSX)

    if SHEET_CHART in wb.sheetnames:
        ws = wb[SHEET_CHART]
        wb.remove(ws)
    ws = wb.create_sheet(SHEET_CHART)

    img = XLImage(str(fig1_path))
    img.anchor = "A1"
    ws.add_image(img)

    wb.save(OUT_XLSX)


def main():
    print("=== Garmin 心拍 → Excel（重複なし） ===")
    start_d = input_date("開始日 (YYYY-MM-DD)", DEFAULT_START)
    end_d = input_date("終了日 (YYYY-MM-DD)", DEFAULT_END)

    if start_d > end_d:
        raise ValueError("開始日が終了日より後です。")

    df_new = fetch_garmin_daily_rows(start_d, end_d)
    df_all = upsert_raw_to_excel(df_new)
    df_month = make_monthly_summary(df_all)

    write_excel(df_all, df_month)

    print("✅ 完了:", OUT_XLSX)
    print(f" - {SHEET_RAW}: 日次データ {len(df_all)} 行")
    print(f" - {SHEET_MONTH}: 月別集計 {len(df_month)} 行")
    print(f" - {SHEET_CHART}: グラフ貼り付け")


if __name__ == "__main__":
    main()