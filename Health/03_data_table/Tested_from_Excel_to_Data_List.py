# %%
#Book「食事記録表(--0240806).xlsx」のデータから体重、血糖値、血圧、中程度運動、運動消費エネルギー、歩数のデータをBook「データ表1」を作成する(20240623確認)

import openpyxl
from openpyxl import utils
from openpyxl import load_workbook

#プログラムとデータはすべてCドライブのPythonWorkに置く

import os

import argparse
import datetime

parser = argparse.ArgumentParser()

parser.add_argument("--start-date", required=True)
parser.add_argument("--end-date", required=True)

args = parser.parse_args()

dt_start_input = datetime.datetime.strptime(
    args.start_date, "%Y-%m-%d"
).date()

dt_end_input = datetime.datetime.strptime(
    args.end_date, "%Y-%m-%d"
).date()

print("開始日", dt_start_input)
print("終了日", dt_end_input)

def get_excel_path(filename, folder="ExcelDATA"):
    base = os.path.join(os.environ["OneDrive"], "ドキュメント", "PythonWork")
    return os.path.join(base, folder, filename)

filepath1 = get_excel_path("食事記録表(--20240806-3).xlsx")
filepath2 = get_excel_path("データ表1.xlsx")

wb1=load_workbook(filename=filepath1)
wb2=load_workbook(filename=filepath2)

ws1=wb1["食事"]

ws5=wb2['Sheet1']
ws6=wb2['Sheet2']
ws7=wb2['Sheet4']

# データ表のリセット
wb2.remove(wb2['Sheet1'])
ws5=wb2.create_sheet('Sheet1')
wb2.remove(wb2['Sheet2'])
ws6=wb2.create_sheet('Sheet2')
wb2.remove(wb2['Sheet4'])
ws7=wb2.create_sheet('Sheet4')

Sub_title1="日付"
Sub_title2="体重"
Sub_title3="血糖値"
Sub_title4="血圧収縮期"
Sub_title5="血圧拡張期"
Sub_title6="心拍数"
Sub_title7="中程度運動量（分）"
Sub_title8="運動消費カロリー"
Sub_title9="歩数"

Column_counter=0

ws5.cell(row=1, column=Column_counter+1).value=Sub_title1
ws5.cell(row=1, column=Column_counter+2).value=Sub_title2
ws5.cell(row=1, column=Column_counter+3).value=Sub_title3
ws5.cell(row=1, column=Column_counter+4).value=Sub_title4
ws5.cell(row=1, column=Column_counter+5).value=Sub_title5
ws5.column_dimensions['A'].width = 12  #Columnの幅を広げる

ws7.cell(row=1, column=Column_counter+1).value=Sub_title1
ws7.cell(row=1, column=Column_counter+2).value=Sub_title2
ws7.cell(row=1, column=Column_counter+3).value=Sub_title3
ws7.cell(row=1, column=Column_counter+4).value=Sub_title4
ws7.cell(row=1, column=Column_counter+5).value=Sub_title5
ws7.cell(row=1, column=Column_counter+6).value=Sub_title6
ws7.cell(row=1, column=Column_counter+7).value=Sub_title7
ws7.cell(row=1, column=Column_counter+8).value=Sub_title8
ws7.cell(row=1, column=Column_counter+9).value=Sub_title9

ws7.column_dimensions['A'].width = 12  #Columnの幅を広げる


# 検索開始日時をExcel Data(食事記録)から取得する
import re

table_Column_num=1
table_Row_num=2

k=0
Hit_mark=0

Pick_Date=ws1.cell(row=2,column=1).value  #first date of the excel file
#Pick_Date1=utils.datetime.from_excel(Pick_Date)  #Pick_Dateを年月日時刻に変形する
Pick_Date2=Pick_Date.date() 
#print(dt_start_input, Pick_Date2)

while Pick_Date2!=dt_start_input:       #Excelの日付データが検索開始日付と等しくない限り続ける
    k=k+1
    if k>200:
        break
    else:
        for i in range(11):      #Excelには11日分が7列で1セットになっている
            Pick_Date=ws1.cell(row=table_Row_num,column=table_Column_num).value  #日付dataをExcelから読み込む
            
            if Pick_Date!=None:
                #Pick_Date1=utils.datetime.from_excel(Pick_Date)  #Pick_Dateを年月日時刻に変形する
                Pick_Date2=Pick_Date.date()                     #年月日だけを取り出す
                # print("k", k, "i", i, Pick_Date2)
                if Pick_Date2==dt_start_input:
                    Hit_mark=1
                    Start_Date=Pick_Date2
                    Start_Row=table_Row_num
                    Start_Column=table_Column_num
                else:
                    table_Row_num=table_Row_num+7
            else:
                print("End of Data")
                break
                
    table_Row_num=2
    table_Column_num=table_Column_num+7

if Hit_mark==0:
    print(k)
    print("S-date is not found")
else:
    #print("Start Date:",Start_Date)
    #print("Start Row:", Start_Row)
    #print("Start Column:", Start_Column)
    #Column_Address=utils.get_column_letter(Start_Column)
    #print("Start Column Address:", Column_Address)

#検索終了日付の取得

    table_Column_num=Start_Column
    table_Row_num=Start_Row
    k=0
    Hit_mark=0
    if table_Row_num==72:
        table_Row_num=2
        table_Column_num=table_Column_num+7

    Pick_Date=ws1.cell(row=table_Row_num,column=table_Column_num).value
    #Pick_Date1=utils.datetime.from_excel(Pick_Date)  #Pick_Dateを年月日時刻に変形する
    Pick_Date2=Pick_Date.date()
    #print("E date:",dt_end_input) 
    #print("SE date", Pick_Date2)

    if Pick_Date==None:
        print("Endof the Date is out of range")
    
    else:
        Pick_Date=ws1.cell(row=table_Row_num,column=table_Column_num).value  #seach date of the excel file
        #Pick_Date1=utils.datetime.from_excel(Pick_Date)  #Pick_Dateを年月日時刻に変形する
        Pick_Date2=Pick_Date.date() 
        #print("開始前の初めの日付", Pick_Date2)
    
        while Pick_Date2!=dt_end_input:       #Excelの日付データが検索終了日付と等しくない限り続ける
            k=k+1
            #print(k)
            if k>200:
                break

            else:
                for i in range(11):      #Excelには11日分が7列で1セットになっている
                    Pick_Date=ws1.cell(row=table_Row_num,column=table_Column_num).value  #日付dataをExcelから読み込む
                    if Pick_Date!=None:
                        #Pick_Date1=utils.datetime.from_excel(Pick_Date)  #Pick_Dateを年月日時刻に変形する
                        Pick_Date2=Pick_Date.date()                     #年月日だけを取り出す
                        #print(i, type(Pick_Date2), Pick_Date2,  type(dt_end_input), dt_end_input)
                        #print(table_Row_num, table_Column_num)
                
                        if Pick_Date2==dt_end_input:
                            Hit_mark=1
                            End_Date=Pick_Date2
                            End_Row=table_Row_num
                            End_Column=table_Column_num
                        else:
                            table_Row_num=table_Row_num+7   
                    else:
                        print("End of Data")
                        break
            table_Row_num=2
            table_Column_num=table_Column_num+7

        if Hit_mark==0:
            print(k)
            print("E date is not found")
        else:
            #print("End Date:",End_Date)
            #print("End Row:", End_Row)
            #print("End Column:", End_Column)
            #Column_Address=utils.get_column_letter(End_Column)
            #print("End Column Address:", Column_Address)
        
            print(Pick_Date2)
            


# 9データの読み取り 2025/01/10

table_Row_num=Start_Row
table_Column_num=Start_Column
Pick_Date2=dt_start_input
Pick_Weight=0
print("Start_Row:", Start_Row, "Start_Column:", Start_Column, "Pick_Date2:", Pick_Date2)
k=0
write_row=1
write_column=1
start_equal=0
start_step=0
initial_Column=0

Hit_Mark=0

while Pick_Date2!=dt_end_input:       #Excelの日付データが検索終了日付と等しくない限り続ける
        k=k+1
        #print(k)
        if k>3000:
            break

        else:
            for i in range(11):      #Excelには11日分が7列で1セットになっている
                if Hit_Mark==1:
                    break
                Pick_Date=ws1.cell(row=table_Row_num,column=table_Column_num).value  #日付dataをExcelから読み込む
                if Pick_Date!=None:
                    #Pick_Date1=utils.datetime.from_excel(Pick_Date)  #Pick_Dateを年月日時刻に変形する
                    Pick_Date2=Pick_Date.date()                     #年月日だけを取り出す
                    #print(k, i, Hit_Mark, Pick_Date2, table_Row_num, table_Column_num)
                    if Pick_Date2==dt_end_input:
                        Hit_Mark=1
                elif Hit_Mark==1:
                    break
                else:
                    break
                
                #　体重
                Pick_Weight1=ws1.cell(row=table_Row_num+3, column=table_Column_num+2).value
                #print(k, Pick_Date2, type(Pick_Weight1), Pick_Weight1)
                if Pick_Weight1!=None:
                    #print("Yes not None", type(Pick_Weight1))
                    if isinstance(Pick_Weight1, float):
                        Pick_Weight=Pick_Weight1
                    elif isinstance(Pick_Weight1, int):
                        Pick_Weight=float(Pick_Weight1)
                    elif Pick_Weight1=="-":
                        Pick_Weight=None
                    elif isinstance(Pick_Weight1, str) and Pick_Weight1[2]==",":
                        Weight1=Pick_Weight1.replace(",", ".")
                        Pick_Weight=float(Weight1)
                else:
                    Pick_Weight=None
                
                #print(k, "Pick_Weight", Pick_Weight)
                
                #血糖値
                Pick_BloodS2=ws1.cell(row=table_Row_num+5, column=table_Column_num+2).value
                if Pick_BloodS2!=None:
                    Pick_BloodS1=Pick_BloodS2.replace(" ", "")
                #if Pick_BloodS1!=None:
                    if Pick_BloodS1[3]=="/":
                        Pick_BloodS=int(Pick_BloodS1[:3])
                    elif Pick_BloodS1[2]=="/":
                        Pick_BloodS=int(Pick_BloodS1[:2])
                else:
                    Pick_BloodS=None
                
                #血圧
                Pick_BloodP2=ws1.cell(row=table_Row_num+4, column=table_Column_num+4).value
                if Pick_BloodP2!=None:
                    Pick_BloodP1=Pick_BloodP2.replace(" ", "")
                    #print("Pick_BloodP1:", Pick_BloodP1)
                #if Pick_BloodP1!=None:
                    if Pick_BloodP1[3]=="-":#収縮期が3桁の場合
                        BloodPH=int(Pick_BloodP1[:3])#収縮期
                        #print("BloodPH:", BloodPH)
                        BloodPL=int(Pick_BloodP1[4:6])#拡張期
                        BloodHeartRate=int(Pick_BloodP1[7:9])#心拍数
                        
                    elif Pick_BloodP1[2]=="-":#収縮期が2桁の場合
                        BloodPH=int(Pick_BloodP1[:2])#収縮期
                        BloodPL=int(Pick_BloodP1[3:5])#拡張期
                        BloodHeartRate=int(Pick_BloodP1[6:8])#心拍数
                else:
                    BloodPH=None
                    BloodPL=None
                    BloodHeartRate=None
                    
                #"中程度運動量（分）"
                MiddleActivity=ws1.cell(row=table_Row_num+3, column=table_Column_num+6).value
                if MiddleActivity!=None:
                    if isinstance(MiddleActivity, int):
                        MiddleActivity=MiddleActivity
                    elif isinstance(MiddleActivity, str):
                        MiddleActivity=int(MiddleActivity)
                else:
                    MiddleActivity=None
                # print("中程度運動量（分）", MiddleActivity, type(MiddleActivity))    
                 #"運動消費カロリー"
                ActivityCal=ws1.cell(row=table_Row_num+5, column=table_Column_num+6).value
                #print("ActivityCal", type(ActivityCal), ActivityCal)
                if ActivityCal!=None:
                    ActivityCal=int(ActivityCal)
                   # print("2ActivityCal", type(ActivityCal), ActivityCal)
                
                 #"歩数"
                Steps=ws1.cell(row=table_Row_num+5, column=table_Column_num+4).value
                if Steps!=None:
                    # if isinstance(Steps, int):
                    #     Steps=Steps
                    # else:
                    #     Steps=int(Steps)
                    # 文字列化して数字だけ抽出
                    s = ''.join(re.findall(r'\d+', str(Steps)))
                    if s:
                        Steps = int(s)
                    else:
                        Steps = None

                else:
                    Steps=None
                    
                    
                print("Date", Pick_Date2)
                #print("Weight", Pick_Weight)
                #print("Blood S", Pick_BloodS)
                #print("Blood P H", BloodPH)
                #print("Blood P L", BloodPL)

                table_Row_num=table_Row_num+7
                
                #print("Initial Column_counter", Column_counter)
#読み取ったデータのExcelへの書き込み
                write_row=write_row+1
                write_column=0
                ws7.cell(row=write_row, column=write_column+1).value=Pick_Date2
                ws7.cell(row=write_row, column=write_column+2).value=Pick_Weight
                ws7.cell(row=write_row, column=write_column+3).value=Pick_BloodS
                ws7.cell(row=write_row, column=write_column+4).value=BloodPH
                ws7.cell(row=write_row, column=write_column+5).value=BloodPL    
                ws7.cell(row=write_row, column=write_column+6).value=BloodHeartRate
                ws7.cell(row=write_row, column=write_column+7).value=MiddleActivity
                ws7.cell(row=write_row, column=write_column+8).value=ActivityCal
                ws7.cell(row=write_row, column=write_column+9).value=Steps
                
            if Hit_Mark==1:
                break

        table_Row_num=2
        table_Column_num=table_Column_num+7

wb2.save(filepath2)

print("データ転送終了")
wb1.close()
wb2.close()

import sys
sys.exit

# %%



