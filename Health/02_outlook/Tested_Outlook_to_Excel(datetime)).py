# Outlook予定表から指定期間のデータを取得し、
# 食事記録表(--20240806-3).xlsx の「食事」シートへ転記する
#
# Power Automate対応版
# 実行例:
# python "Tested_Outlook_to_Excel(datetime)).py" --start-date 2026-06-15 --end-date 2026-06-25

import os
import re
import argparse
import datetime as dt
import unicodedata

import win32com.client as win32
import openpyxl
from openpyxl import load_workbook, utils
from openpyxl.styles import Alignment, PatternFill, Font


# =========================
# 引数取得
# =========================

parser = argparse.ArgumentParser()
parser.add_argument("--start-date", required=True, help="開始日 YYYY-MM-DD")
parser.add_argument("--end-date", required=True, help="終了日 YYYY-MM-DD")
args = parser.parse_args()

START_DATE = dt.datetime.strptime(args.start_date, "%Y-%m-%d").date()
END_DATE = dt.datetime.strptime(args.end_date, "%Y-%m-%d").date()

if END_DATE < START_DATE:
    raise ValueError("終了日が開始日より前です")

print("開始日:", START_DATE)
print("終了日:", END_DATE)


# =========================
# パス設定
# =========================

def get_excel_path(filename, folder="ExcelDATA"):
    base = os.path.join(os.environ["OneDrive"], "ドキュメント", "PythonWork")
    return os.path.join(base, folder, filename)


EXCEL_PATH = get_excel_path("食事記録表(--20240806-3).xlsx")


# =========================
# Outlook予定表取得
# =========================

PROFILE_NAME = None
STORE_HINTS = ["kishi_isao@outlook.jp"]


def iter_stores(session):
    stores = session.Stores
    for i in range(1, stores.Count + 1):
        yield stores.Item(i)


def pick_store_by_hints(session, hints):
    hints_l = [h.lower() for h in hints]
    for s in iter_stores(session):
        name = (s.DisplayName or "").lower()
        if any(h in name for h in hints_l):
            return s
    return None


def get_calendar_items(start_date, end_date):
    start_dt = dt.datetime.combine(start_date, dt.time(0, 0, 0))
    end_dt = dt.datetime.combine(end_date, dt.time(23, 59, 59))

    app = win32.gencache.EnsureDispatch("Outlook.Application")
    session = app.GetNamespace("MAPI")
    session.Logon(PROFILE_NAME or "", "", False, False)

    print("=== Stores ===")
    for s in iter_stores(session):
        try:
            print("-", s.DisplayName)
        except Exception:
            pass

    store = pick_store_by_hints(session, STORE_HINTS)
    if store is None:
        store = session.DefaultStore

    print("使用ストア:", store.DisplayName)

    try:
        calendar = store.GetDefaultFolder(9)
    except Exception:
        root = store.GetRootFolder()
        for cand in ("予定表", "Calendar", "カレンダー"):
            try:
                calendar = root.Folders[cand]
                break
            except Exception:
                pass
        else:
            raise RuntimeError("予定表フォルダーが見つかりませんでした。")

    items = calendar.Items
    items.IncludeRecurrences = True
    items.Sort("[Start]")

    restriction = (
        f"[Start] >= '{start_dt:%m/%d/%Y %I:%M %p}' AND "
        f"[Start] <= '{end_dt:%m/%d/%Y %I:%M %p}'"
    )

    filtered = items.Restrict(restriction)

    result = []
    for it in filtered:
        try:
            _ = it.Start
            result.append(it)
        except Exception:
            pass

    print("Outlook取得件数:", len(result))
    return result


# =========================
# Excel上の日付位置を探す
# =========================

def excel_value_to_date(value):
    if value is None:
        return None

    if isinstance(value, dt.datetime):
        return value.date()

    if isinstance(value, dt.date):
        return value

    if isinstance(value, int):
        try:
            return utils.datetime.from_excel(value).date()
        except Exception:
            return None

    return None


def find_date_cell(ws, target_date):
    """
    食事シートは、7列単位・7行単位で日付ブロックが並ぶ前提。
    日付セルを探して、Start_Row / Start_Column を返す。
    """

    for col in range(1, ws.max_column + 1, 7):
        for row in range(2, ws.max_row + 1, 7):
            value = ws.cell(row=row, column=col).value
            d = excel_value_to_date(value)

            if d == target_date:
                return row, col

    return None, None


# =========================
# 補助関数
# =========================

def duration_to_text(delta):
    total_minutes = int(delta.total_seconds() // 60)
    hours = total_minutes // 60
    minutes = total_minutes % 60

    if minutes == 0:
        return f"{hours}時間"
    if minutes == 15:
        return f"{hours}.25時間"
    if minutes == 30:
        return f"{hours}.5時間"
    if minutes == 45:
        return f"{hours}.75時間"

    return f"{hours}時間{minutes:02d}分"


def set_activity_cell(ws, row, col, text):
    ws.cell(row=row, column=col).value = text
    ws.cell(row=row, column=col).font = Font(color="008B8B")


# =========================
# 1日分をExcelへ転記
# =========================

def write_one_day(ws, target_date, select_items):
    start_row, start_col = find_date_cell(ws, target_date)

    if start_row is None:
        print("日付がExcel上に見つかりません:", target_date)
        return

    print("転記日:", target_date, "Row:", start_row, "Column:", start_col)

    for item in select_items:
        try:
            item_date = item.Start.date()
            subject_raw = str(item.Subject or "")
            subject = unicodedata.normalize("NFKC", subject_raw).strip()
        except Exception:
            continue

        if item_date != target_date:
            continue

        # 朝食・昼食・夕食
        if "朝食　" in subject_raw:
            ws.cell(row=start_row, column=start_col + 2).value = subject_raw[3:]

        if "昼食　" in subject_raw:
            ws.cell(row=start_row + 1, column=start_col + 2).value = subject_raw[3:]

        if "夕食　" in subject_raw:
            ws.cell(row=start_row + 2, column=start_col + 2).value = subject_raw[3:]

        # HA1c
        if "HA1c " in subject_raw:
            ha1c = subject_raw[5:]
            ws.cell(row=start_row + 6, column=start_col + 2).value = "定期健診　HA1c"
            ws.cell(row=start_row + 6, column=start_col + 4).value = ha1c
            ws.cell(row=start_row + 6, column=start_col + 2).font = Font(color="FF0000")
            ws.cell(row=start_row + 6, column=start_col + 4).font = Font(color="FF0000")

        # 服薬
        if "服薬" in subject:
            fukuyaku = "○" if any(mark in subject for mark in ("〇", "○", "◯")) else "×"
            cell = ws.cell(row=start_row + 4, column=start_col + 2)
            cell.value = fukuyaku
            cell.alignment = Alignment(horizontal="center")

        # 飲酒
        if "飲酒" in subject_raw:
            cell = ws.cell(row=start_row + 4, column=start_col + 6)

            if "◎" in subject_raw:
                cell.value = "◎"
                cell.fill = PatternFill(patternType="solid", fgColor="ADD8E6", bgColor="ADD8E6")
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = subject_raw[2:]
                cell.fill = PatternFill(patternType="solid", fgColor="FFFFFF", bgColor="FFFFFF")

        # 中程度運動 分
        if re.fullmatch(r"\d+分", subject):
            undou = int(subject.replace("分", ""))
            ws.cell(row=start_row + 3, column=start_col + 6).value = undou

        # 運動消費 kcal
        if "運動消費" in subject:
            m = re.search(r"運動消費\s*([0-9]+)\s*kcal", subject)
            if m:
                ws.cell(row=start_row + 5, column=start_col + 6).value = int(m.group(1))

        # 歩数
        if "歩数" in subject:
            m = re.search(r"歩数\s*([0-9]+)", subject)
            if m:
                ws.cell(row=start_row + 5, column=start_col + 4).value = int(m.group(1))

        # 農作業
        if "農作業" in subject:
            delta_text = duration_to_text(item.End - item.Start)

            if item.Start.strftime("%p") == "AM":
                set_activity_cell(ws, start_row + 6, start_col + 3, "農作業" + delta_text)
            else:
                set_activity_cell(ws, start_row + 6, start_col + 4, "農作業" + delta_text)

        # ウォーキング
        if "ウォーキング" in subject:
            walking = subject_raw[7:]
            ws.cell(row=start_row + 3, column=start_col + 4).value = walking

        # ゴルフ練習
        golf_exercise = False

        if "ゴルフ練習" in subject:
            golf_exercise = True
            delta_text = duration_to_text(item.End - item.Start)

            if item.Start.strftime("%p") == "AM":
                set_activity_cell(ws, start_row + 6, start_col + 3, "ゴルフ練習" + delta_text)
            else:
                set_activity_cell(ws, start_row + 6, start_col + 4, "ゴルフ練習" + delta_text)

        # ゴルフ
        if "ゴルフ" in subject and not golf_exercise:
            golf_play = "ゴルフ" + str(item.Location or "")
            set_activity_cell(ws, start_row + 6, start_col + 3, golf_play)

        # 朝食前：体重・血糖値・血圧
        if "朝食前" in subject:
            # 体重
            m = re.search(r"体重\s*([0-9]+(?:\.[0-9]+)?)\s*kg", subject)
            taijyu = float(m.group(1)) if m else None
            ws.cell(row=start_row + 3, column=start_col + 2).value = taijyu

            # 血糖値
            m = re.search(r"血糖値[^0-9]*([0-9]{2,3}/[0-9]{1,2}:[0-9]{2})", subject)
            kettouti = m.group(1) if m else None

            cell = ws.cell(row=start_row + 5, column=start_col + 2)
            cell.value = kettouti

            cell.fill = PatternFill(patternType=None)
            cell.font = Font(color="000000")

            if kettouti:
                value = int(kettouti.split("/")[0])

                if value >= 130:
                    cell.font = Font(color="FF0000")

                elif value < 120:
                    cell.font = Font(color="000000")
                    cell.fill = PatternFill(patternType="solid", fgColor="ADFF2F")

            # 血圧
            bp_list = re.findall(
                r"[0-9]{2,3}-[0-9]{2,3}/[0-9]{1,4}/[0-9]{1,2}:[0-9]{2}",
                subject,
            )

            ketuatu = bp_list[-1].strip(", ").strip() if bp_list else None
            ws.cell(row=start_row + 4, column=start_col + 4).value = ketuatu

            if ketuatu:
                m = re.match(r"([0-9]{2,3})-[0-9]{2,3}/[0-9]{1,4}/[0-9]{1,2}:[0-9]{2}", ketuatu)
                if m:
                    sys_bp = int(m.group(1))
                    if sys_bp <= 50:
                        alert_date = item.Start.strftime("%Y-%m-%d %H:%M")
                        ws.cell(row=start_row + 4, column=start_col + 5).value = f"異常({sys_bp}) {alert_date}"
                        print(f"⚠ 異常血圧: {sys_bp} {alert_date}")


# =========================
# メイン処理
# =========================

def main():
    select_items = get_calendar_items(START_DATE, END_DATE)

    wb = load_workbook(filename=EXCEL_PATH)
    ws = wb["食事"]

    current_date = START_DATE

    while current_date <= END_DATE:
        write_one_day(ws, current_date, select_items)
        current_date += dt.timedelta(days=1)

    wb.save(EXCEL_PATH)
    wb.close()

    print("=== Outlook → Excel 転記完了 ===")
    print(EXCEL_PATH)


if __name__ == "__main__":
    main()



