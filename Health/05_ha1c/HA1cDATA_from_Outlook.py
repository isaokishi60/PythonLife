# HA1cのデータをOutlookから取得する
# 20240623確認済

# 日付入力しOutlookの予定表(kishi_isao@outlook.jp)からその日のデータを取得する(20251113確認)

#予定表データをOutlookから取得する(時間がかかるので先に実施しておく)

import win32com.client
import datetime
import time
import tkinter as tk

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# ★アカウント名を直接指定してルートフォルダを取得
root_folder = outlook.Folders["kishi_isao@outlook.jp"]

# 予定表フォルダを取得
calendar = root_folder.Folders["予定表"]

items = calendar.Items
items.Sort("[Start]")

select_items = []  # 指定した期間内の予定を入れるリスト

# 予定を抜き出す期間
start_date = datetime.date(2020, 1, 1)
end_date = datetime.date(2026, 7, 31)


# --------------------------------------------------
# 進行状況を表示するウィンドウ
# --------------------------------------------------
progress_window = tk.Tk()
progress_window.title("HA1cデータ取得中")
progress_window.geometry("460x170")
progress_window.resizable(False, False)

title_label = tk.Label(
    progress_window,
    text="Outlookの予定表を確認しています",
    font=("Yu Gothic UI", 14, "bold")
)
title_label.pack(pady=(20, 10))

progress_label = tk.Label(
    progress_window,
    text="処理を開始しています...",
    font=("Yu Gothic UI", 11),
    justify="left"
)
progress_label.pack(pady=5)

notice_label = tk.Label(
    progress_window,
    text="この画面が表示されている間は処理中です。",
    font=("Yu Gothic UI", 9)
)
notice_label.pack(pady=5)

progress_window.update()

# 開始時刻
process_start_time = time.monotonic()
last_update_time = process_start_time

processed_count = 0
selected_count = 0


try:
    for item in items:

        processed_count += 1

        try:
            item_date = item.Start.date()

            if start_date <= item_date <= end_date:
                select_items.append(item)
                selected_count += 1

        except Exception:
            # 開始日時を取得できない項目は飛ばす
            continue

        # 0.5秒ごとに画面を更新
        current_time = time.monotonic()

        if current_time - last_update_time >= 0.5:

            elapsed_seconds = int(current_time - process_start_time)
            elapsed_minutes, elapsed_seconds = divmod(elapsed_seconds, 60)

            progress_label.config(
                text=(
                    f"確認した予定：{processed_count:,} 件\n"
                    f"対象期間内　：{selected_count:,} 件\n"
                    f"経過時間　　：{elapsed_minutes}分 {elapsed_seconds:02d}秒"
                )
            )

            progress_window.update_idletasks()
            progress_window.update()

            last_update_time = current_time


    # 最終表示
    total_seconds = int(time.monotonic() - process_start_time)
    total_minutes, total_seconds = divmod(total_seconds, 60)

    progress_label.config(
        text=(
            f"Outlook予定表の確認が完了しました。\n"
            f"確認した予定：{processed_count:,} 件\n"
            f"取得した予定：{selected_count:,} 件\n"
            f"処理時間　　：{total_minutes}分 {total_seconds:02d}秒"
        )
    )

    title_label.config(text="予定表の確認完了")
    notice_label.config(text="続いてデータの書き出し処理を行います。")

    progress_window.update_idletasks()
    progress_window.update()

    # 完了表示を1.5秒間表示
    time.sleep(1.5)

finally:
    progress_window.destroy()

# %%
# HA1cが含まれていれば、日付とデータをC:\Users\spax2\Documents\PythonWork\ExcelDATA\データ表4.xlsxのSheet2に入れるする
import openpyxl
from openpyxl import utils
from openpyxl import load_workbook

import os

def get_excel_path(filename, folder="ExcelDATA"):
    base = os.path.join(os.environ["OneDrive"], "ドキュメント", "PythonWork")
    return os.path.join(base, folder, filename)

filepath = get_excel_path("データ表4.xlsx", folder="ExcelDATA")

wb = load_workbook(filename=filepath)

# --- Sheet2 が存在すれば削除 ---
if "Sheet2" in wb.sheetnames:
    wb.remove(wb["Sheet2"])

# --- 新しい Sheet2 を作成 ---
ws = wb.create_sheet("Sheet2")

# --- ヘッダーを書き込み ---
ws.cell(row=1, column=1).value = "No."
ws.cell(row=1, column=2).value = "日付"
ws.cell(row=1, column=3).value = "HA1c(%)"

Column_Address=utils.get_column_letter(1)  #列番号をアルファベットに変更
ws.column_dimensions[Column_Address].width=4                #列幅を狭くする
Column_Address=utils.get_column_letter(2)  #列番号をアルファベットに変更
ws.column_dimensions[Column_Address].width=12                #列幅を広げる

#print(datetime.date(2023,8,3))

k = 0

for select_item in select_items:

    # 件名を取得（None対策）
    subject = select_item.Subject or ""

    # HA1cを含まないものは飛ばす
    if "HA1c" not in subject:
        continue

    try:
        # HA1cの数値を取得
        hba1c_value = float(subject[5:8])

    except ValueError:
        print("HA1cを数値に変換できません。")
        print(select_item.Start.date())
        print(subject)
        continue

    except Exception as e:
        print("その他のエラー")
        print(select_item.Start.date())
        print(subject)
        print(e)
        continue

    # 正常に取得できた場合のみ書き込む
    k += 1

    print(k)
    print(select_item.Start.date())
    print("HA1c:", hba1c_value)

    ws.cell(row=k + 1, column=1).value = k
    ws.cell(row=k + 1, column=2).value = select_item.Start.date()
    ws.cell(row=k + 1, column=3).value = hba1c_value

# Sheet2をアクティブにする
wb.active = wb["Sheet2"]

# 保存だけ行う
wb.save(filepath)




