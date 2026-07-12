# %%
# 血圧以外の場合

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import numbers


def export_monthly_table(input_path, output_path, sheet_name, item_name):
    rounding_rules = {
        "体重": 1,
        "血糖値": 0,  # データ本体は整数、平均・標準偏差のみ後で調整
        "血圧収縮期": 0,
        "血圧拡張期": 0,
        "心拍数": 0,
        "中程度運動量（分）": 0,
        "運動消費カロリー": 0,
        "歩数": 0
    }

    # 元データ読み込み
    df = pd.read_excel(input_path, sheet_name="Sheet4")
    df["日付"] = pd.to_datetime(df["日付"])
    df = df[["日付", item_name]].dropna()

    df["月"] = df["日付"].dt.month
    df["日"] = df["日付"].dt.day
    df["年"] = df["日付"].dt.year

    # ★ 年 + 月 を列にしたピボットテーブル
    pivot = df.pivot_table(
        index="日",
        columns=["年", "月"],
        values=item_name,
        aggfunc="mean"
    )
    pivot.index = pivot.index.astype(str)

    # 平均・標準偏差の計算
    avg_row = pivot.mean().round(2)
    std_row = pivot.std().round(2)

    # 既存ファイルを開く（なければ新規作成）
    try:
        wb = load_workbook(output_path)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)

    # 対象シートだけ削除して再作成
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # ヘッダー（A1に項目名、B1以降に「2024年1月」「2024年2月」…）
    ws.cell(row=1, column=1, value=item_name)
    for col_index, (year, month) in enumerate(pivot.columns, start=2):
        header = f"{year}年{month}月"
        ws.cell(row=1, column=col_index, value=header)


    # データ行（1〜31日分を固定）
    MAX_DAY_ROWS = 31

    # 行ラベル（日）
    for row_index in range(2, 2 + MAX_DAY_ROWS):
        day_str = str(row_index - 1)
        cell_day = ws.cell(row=row_index, column=1, value=f"{day_str}日")
        cell_day.alignment = Alignment(horizontal="right")

        # 各 年・月 列に書き込み
        for col_index, (year, month) in enumerate(pivot.columns, start=2):
            if day_str in pivot.index:
                value = pivot.loc[day_str, (year, month)]
            else:
                value = None

            if value is not None:
                digits = rounding_rules.get(item_name, 0)
                value = round(value, digits)

            cell_value = ws.cell(row=row_index, column=col_index, value=value)

            # 体重だけ小数1桁表示
            if item_name == "体重" and value is not None:
                cell_value.number_format = "0.0"

    # 平均・標準偏差（固定行）
    avg_row = pivot.mean().round(2)  # MultiIndex の Series
    std_row = pivot.std().round(2)

    ws.cell(row=2 + MAX_DAY_ROWS, column=1, value="AVE")
    ws.cell(row=3 + MAX_DAY_ROWS, column=1, value="STD")
    for col_index, key in enumerate(pivot.columns, start=2):  # key = (year, month)
        ws.cell(row=2 + MAX_DAY_ROWS, column=col_index, value=avg_row[key])
        ws.cell(row=3 + MAX_DAY_ROWS, column=col_index, value=std_row[key])


    # 列幅調整
    for col_index, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = max_length + 2

    # 罫線スタイル（下線のみ）
    thin_border = Border(bottom=Side(style='thin'))

    # 対象行に罫線を引く（1行目、32〜34行目）
    for row in [1, 32, 33, 34]:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = thin_border


    wb.save(output_path)



# %%
# 血圧の場合

def export_grouped_items_to_sheet(input_path, output_path, sheet_name, item_list):
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side
    from openpyxl.styles import Alignment
    from openpyxl.styles import numbers

    rounding_rules = {
        "体重": 1,
        "血糖値": 0,  # データ本体は整数、平均・標準偏差のみ後で調整
        "血圧収縮期": 0,
        "血圧拡張期": 0,
        "心拍数": 0,
        "中程度運動量（分）": 0,
        "運動消費カロリー": 0,
        "歩数": 0
    }

    # 元データ読み込み（※入力シート名は従来どおり Sheet4 固定）
    df = pd.read_excel(input_path, sheet_name="Sheet4")
    df["日付"] = pd.to_datetime(df["日付"])
    df["年"] = df["日付"].dt.year
    df["月"] = df["日付"].dt.month
    df["日"] = df["日付"].dt.day

    # 既存ファイルを開く（なければ新規作成）
    try:
        wb = load_workbook(output_path)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)

    # 対象シートを削除して再作成
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # ★ (年, 月) の組み合わせごとに処理する
    year_month_list = (
        df[["年", "月"]]
        .drop_duplicates()
        .sort_values(["年", "月"])
        .itertuples(index=False, name=None)  # → (year, month) のタプルの列挙
    )

    start_col = 1
    MAX_DAY_ROWS = 31  # 最大日数（1〜31日）

    for year, month in year_month_list:
        # 対象の 年・月 のデータだけ抽出
        df_month = df[(df["年"] == year) & (df["月"] == month)]
        if df_month.empty:
            continue

        # ピボットテーブル（行：日、列：項目）
        table = df_month.pivot_table(index="日", values=item_list, aggfunc="mean")
        table.index = table.index.astype(str)

        # 平均・標準偏差の計算
        avg_row = table.mean().round(2)
        std_row = table.std().round(2)

        # ヘッダー（1行目：見出し、2行目：項目名）
        ws.cell(row=1, column=start_col, value=f"{year}年{month}月")
        for i, item in enumerate(["日付"] + item_list):
            ws.cell(row=2, column=start_col + i, value=item)

        # 日付とデータ（1〜31日分）
        for row_index in range(3, 3 + MAX_DAY_ROWS):
            day_str = str(row_index - 2)  # "1"〜"31"

            # 日付表示
            cell_day = ws.cell(row=row_index, column=start_col, value=f"{day_str}日")
            cell_day.alignment = Alignment(horizontal="right")

            # 各項目の値
            for col_index, item in enumerate(item_list, start=1):
                # table の index にその日があれば値を取得、なければ None
                value = table[item].get(day_str, None) if day_str in table.index else None
                if value is not None:
                    digits = rounding_rules.get(item, 0)
                    value = round(value, digits)

                cell_val = ws.cell(row=row_index, column=start_col + col_index, value=value)

        # 平均・標準偏差（固定行）
        ave_row_idx = 3 + MAX_DAY_ROWS     # 3 + 31 = 34
        std_row_idx = 4 + MAX_DAY_ROWS     # 35

        ws.cell(row=ave_row_idx, column=start_col, value="AVE")
        ws.cell(row=std_row_idx, column=start_col, value="STD")
        for col_index, item in enumerate(item_list, start=1):
            ws.cell(row=ave_row_idx, column=start_col + col_index, value=avg_row.get(item))
            ws.cell(row=std_row_idx, column=start_col + col_index, value=std_row.get(item))

        # 次の (年, 月) ブロックの開始列（項目数+1列分あける）
        start_col += len(item_list) + 2

    # 列幅調整
    for col_index, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_length = 0
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = max_length + 2

    # 罫線スタイル（下線のみ）
    thin_border = Border(bottom=Side(style='thin'))

    # 対象行に罫線を引く（1行目、2行目、最終行側のAVE/STD行）
    # AVE/STD行は 34,35 行目前提（MAX_DAY_ROWS=31 のとき）
    for row in [1, 2, 33, 34, 35]:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = thin_border

    wb.save(output_path)

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from pathlib import Path

plt.rcParams["font.family"] = ["Yu Gothic", "Meiryo", "sans-serif"]
plt.rcParams["axes.unicode_minus"] = False

def create_monthly_mean_std_graphs(input_path, output_dir, item_definitions):
    """
    Sheet4のデータから、各項目の月平均と平均±1σグラフを作成する。

    item_definitions:
        {
            "元データの列名": {
                "title": "グラフタイトル",
                "ylabel": "縦軸名",
                "filename": "保存ファイル名.png"
            }
        }
    """

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 元データを一度だけ読み込む
    df = pd.read_excel(input_path, sheet_name="Sheet4")
    df["日付"] = pd.to_datetime(df["日付"], errors="coerce")
    df = df.dropna(subset=["日付"])

    # 年月単位の集計用列
    df["年月"] = df["日付"].dt.to_period("M")

    for item_name, graph_info in item_definitions.items():

        if item_name not in df.columns:
            print(f"列がないためグラフを作成しません: {item_name}")
            continue

        item_df = df[["年月", item_name]].copy()
        item_df[item_name] = pd.to_numeric(
            item_df[item_name],
            errors="coerce"
        )
        item_df = item_df.dropna(subset=[item_name])

        if item_df.empty:
            print(f"データがないためグラフを作成しません: {item_name}")
            continue

        # 現在のpivot.std()と同じ標本標準偏差（ddof=1）
        monthly_stats = (
            item_df
            .groupby("年月")[item_name]
            .agg(["mean", "std"])
            .sort_index()
        )

        # PeriodIndexをグラフ用の日付へ変換
        x = monthly_stats.index.to_timestamp()

        mean_values = monthly_stats["mean"]
        std_values = monthly_stats["std"]

        lower_values = mean_values - std_values
        upper_values = mean_values + std_values

        fig, ax = plt.subplots(figsize=(12, 6))

        # 平均値の折れ線
        line, = ax.plot(
            x,
            mean_values,
            marker="o",
            linewidth=2,
            label="平均値"
        )

        # 平均値±1σの帯
        ax.fill_between(
            x,
            lower_values,
            upper_values,
            alpha=0.2,
            color=line.get_color(),
            label="平均値 ± 1σ"
        )

        ax.set_title(
            f"{graph_info['title']} 月次平均値と±1σ",
            fontsize=16,
            fontweight="bold"
        )
        ax.set_xlabel("年月")
        ax.set_ylabel(graph_info["ylabel"])

        ax.grid(True, alpha=0.3)
        ax.legend()

        # 横軸の日付表示
        locator = mdates.AutoDateLocator(
            minticks=6,
            maxticks=15
        )
        ax.xaxis.set_major_locator(locator)
        ax.xaxis.set_major_formatter(
            mdates.DateFormatter("%Y/%m")
        )

        plt.setp(
            ax.get_xticklabels(),
            rotation=45,
            ha="right"
        )

        fig.tight_layout()

        png_path = output_dir / graph_info["filename"]

        fig.savefig(
            png_path,
            dpi=180,
            bbox_inches="tight"
        )
        plt.close(fig)

        print("月次グラフ保存:", png_path)


# %%
# データ表作成
import os

def get_excel_path(filename, folder="ExcelDATA"):
    base = os.path.join(os.environ["OneDrive"], "ドキュメント", "PythonWork")
    return os.path.join(base, folder, filename)

input_path = get_excel_path("データ表1.xlsx")
output_path= get_excel_path("データ表1-1.xlsx")

monthly_graph_dir = os.path.join(
    os.environ["OneDrive"],
    "ドキュメント",
    "PythonWork",
    "Health",
    "monthly_graph_png"
)

MONTHLY_GRAPH_ITEMS = {
    "体重": {
        "title": "体重",
        "ylabel": "体重（kg）",
        "filename": "1_体重_月次平均.png",
    },
    "血糖値": {
        "title": "血糖値",
        "ylabel": "血糖値",
        "filename": "2_血糖値_月次平均.png",
    },
    "血圧収縮期": {
        "title": "血圧収縮期",
        "ylabel": "収縮期血圧",
        "filename": "3_血圧収縮期_月次平均.png",
    },
    "血圧拡張期": {
        "title": "血圧拡張期",
        "ylabel": "拡張期血圧",
        "filename": "4_血圧拡張期_月次平均.png",
    },
    "心拍数": {
        "title": "心拍数",
        "ylabel": "心拍数",
        "filename": "5_心拍数_月次平均.png",
    },
    "中程度運動量（分）": {
        "title": "中程度運動量",
        "ylabel": "中程度運動量（分）",
        "filename": "6_中程度運動量_月次平均.png",
    },
    "運動消費カロリー": {
        "title": "運動消費エネルギー",
        "ylabel": "運動消費カロリー（kcal）",
        "filename": "7_運動消費エネルギー_月次平均.png",
    },
    "歩数": {
        "title": "歩数",
        "ylabel": "歩数",
        "filename": "8_歩数_月次平均.png",
    },
}

export_monthly_table(
    input_path,
    output_path,
    sheet_name="体重",
    item_name="体重"
)


export_monthly_table(
    input_path,
    output_path,
    sheet_name="血糖値",
    item_name="血糖値"
)
export_grouped_items_to_sheet(
    input_path,
    output_path,
    sheet_name="血圧",
    item_list=["血圧収縮期", "血圧拡張期", "心拍数"]
)
export_monthly_table(
    input_path,
    output_path,
    sheet_name="中程度運動量",
    item_name="中程度運動量（分）"
)
export_monthly_table(
    input_path,
    output_path,
    sheet_name="運動消費カロリー",
    item_name="運動消費カロリー"
)
export_monthly_table(
    input_path,
    output_path,
    sheet_name="歩数",
    item_name="歩数"
)


create_monthly_mean_std_graphs(
    input_path=input_path,
    output_dir=monthly_graph_dir,
    item_definitions=MONTHLY_GRAPH_ITEMS
)


print(f"出力ファイルを開きます: {output_path}")

# Excelで開く（Pythonは待たずに終了）
os.startfile(output_path)



