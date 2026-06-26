# %%
# 9個のデータの月平均値のグラフを作成する　20250611確認済

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from datetime import datetime
from collections import defaultdict

# Excelファイルの読み込み
import os

def get_excel_path(filename, folder="ExcelDATA"):
    base = os.path.join(os.environ["OneDrive"], "ドキュメント", "PythonWork")
    return os.path.join(base, folder, filename)

filepath= get_excel_path("データ表1.xlsx", folder="ExcelDATA")


wb = openpyxl.load_workbook(filepath)
ws4 = wb['Sheet4']

# データの格納用
monthly_data = defaultdict(lambda: {
    '体重': [], '血糖値': [], '血圧収縮期': [], '血圧拡張期': [],
    '心拍数': [], '中程度運動量（分）': [], '運動消費カロリー': [], '歩数': []
})

# データ処理（欠損は除いて処理）
for row in ws4.iter_rows(min_row=2, values_only=True):
    date_str, weight, glucose, bph, bpl, hb, active_min, active_cal, steps = row

    if not date_str:
        continue  # 日付がない行はスキップ

    date = date_str if isinstance(date_str, datetime) else datetime.strptime(date_str, '%Y/%m/%d')
    month_key = date.strftime('%Y-%m')

    if weight is not None:
        monthly_data[month_key]['体重'].append(weight)
    if glucose is not None:
        monthly_data[month_key]['血糖値'].append(glucose)
    if bph is not None:
        monthly_data[month_key]['血圧収縮期'].append(bph)
    if bpl is not None:
        monthly_data[month_key]['血圧拡張期'].append(bpl)
    if hb is not None:
        monthly_data[month_key]['心拍数'].append(hb)
    if active_min is not None:
        monthly_data[month_key]['中程度運動量（分）'].append(active_min)
    if active_cal is not None:
        monthly_data[month_key]['運動消費カロリー'].append(active_cal)
    if steps is not None:
        monthly_data[month_key]['歩数'].append(steps)

# 月平均シート作成
if '月平均' in wb.sheetnames:
    del wb['月平均']
summary_ws4 = wb.create_sheet('月平均')

# ヘッダー
headers = ['月', '体重平均', '血糖値平均', '血圧収縮期平均', '血圧拡張期平均',
           '心拍数平均', '中程度運動量（分）平均', '運動消費カロリー平均', '歩数平均']
summary_ws4.append(headers)

# 平均値をシートに書き込み
for month in sorted(monthly_data.keys()):
    data = monthly_data[month]
    averages = []
    for key in ['体重', '血糖値', '血圧収縮期', '血圧拡張期', '心拍数',
                '中程度運動量（分）', '運動消費カロリー', '歩数']:
        values = data[key]
        avg = sum(values) / len(values) if values else ''
        averages.append(round(avg, 2) if avg != '' else '')
    summary_ws4.append([month] + averages)

# 折れ線グラフの作成
# === 各項目ごとに別グラフを作成 ===

# グラフの作成関数
def create_line_chart(title, y_title, data_range, cats_range):
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "月"
    chart.style = 2
    chart.add_data(data_range, titles_from_data=True)
    chart.set_categories(cats_range)
    return chart

# 月の範囲（カテゴリ軸）
cats = Reference(summary_ws4, min_col=1, min_row=2, max_row=summary_ws4.max_row)

# グラフ1：体重
data_w = Reference(summary_ws4, min_col=2, max_col=2, min_row=1, max_row=summary_ws4.max_row)
chart_w = create_line_chart("体重の月平均", "kg", data_w, cats)
summary_ws4.add_chart(chart_w, "K2")

# グラフ2：血糖値
data_g = Reference(summary_ws4, min_col=3, max_col=3, min_row=1, max_row=summary_ws4.max_row)
chart_g = create_line_chart("血糖値の月平均", "mg/dL", data_g, cats)
summary_ws4.add_chart(chart_g, "K20")

# グラフ3：中程度運動量（分）
data_m = Reference(summary_ws4, min_col=7, max_col=7, min_row=1, max_row=summary_ws4.max_row)
chart_m = create_line_chart("中程度運動量（分）の月平均", "分", data_m, cats)
summary_ws4.add_chart(chart_m, "K38")

# グラフ4：運動消費カロリー
data_c = Reference(summary_ws4, min_col=8, max_col=8, min_row=1, max_row=summary_ws4.max_row)
chart_c = create_line_chart("運動消費カロリーの月平均", "kcal", data_c, cats)
summary_ws4.add_chart(chart_c, "K56")

# グラフ5：歩数
data_s = Reference(summary_ws4, min_col=9, max_col=9, min_row=1, max_row=summary_ws4.max_row)
chart_s = create_line_chart("歩数の月平均", "歩", data_s, cats)
summary_ws4.add_chart(chart_s, "K74")

# グラフ6：血圧（高・低）＋心拍数
data_bp_h_l_hb = Reference(summary_ws4, min_col=4, max_col=6, min_row=1, max_row=summary_ws4.max_row)
chart_bp_hl_hb = create_line_chart("血圧と心拍数の月平均", "値", data_bp_h_l_hb, cats)
summary_ws4.add_chart(chart_bp_hl_hb, "K92")

# ファイル保存
wb.save(filepath)


