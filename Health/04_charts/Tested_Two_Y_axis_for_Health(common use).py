# %%
#データ表2のSheet1から指定された期間のデータ（体重、血糖値、血圧、中程度運動、運動消費エネルギー、歩数、）をSheet2に移し
# 選択された2項目を2軸のグラフにしてSheet5に作成する(2025/01/25確認)

import openpyxl
import datetime
from openpyxl import utils
from openpyxl import load_workbook

import os

def get_excel_path(filename, folder="ExcelDATA"):
    base = os.path.join(os.environ["OneDrive"], "ドキュメント", "PythonWork")
    return os.path.join(base, folder, filename)

filepath= get_excel_path("データ表1.xlsx", folder="ExcelDATA")


wb=load_workbook(filename=filepath)
ws1=wb['Sheet1']
ws2=wb['Sheet2']
ws3=wb['Sheet3']
ws4=wb['Sheet4']
ws5=wb['Sheet5']

# データ表Sheet2, Sheet3, Sheet5をのリセット
wb.remove(wb['Sheet2'])
ws2=wb.create_sheet('Sheet2')
wb.remove(wb['Sheet3'])
ws3=wb.create_sheet('Sheet3')
wb.remove(wb['Sheet5'])
ws5=wb.create_sheet('Sheet5')


#Sheet2のrow1に項目名を入れる
Sub_title1="日付"
Sub_title2="体重"
Sub_title3="血糖値"
Sub_title4="血圧収縮期"
Sub_title5="血圧拡張期"
Sub_title6="心拍数"
Sub_title7="中程度運動(分）"
Sub_title8="運動消費エネルギー（Kcal）"
Sub_title9="歩数"

Column_counter=0

# H:/Python/ExcelDATA/データ表1.xlsxのSheet2に項目名を張り付ける

ws2.cell(row=1, column=Column_counter+1).value=Sub_title1
ws2.cell(row=1, column=Column_counter+2).value=Sub_title2
ws2.cell(row=1, column=Column_counter+3).value=Sub_title3
ws2.cell(row=1, column=Column_counter+4).value=Sub_title4
ws2.cell(row=1, column=Column_counter+5).value=Sub_title5
ws2.cell(row=1, column=Column_counter+6).value=Sub_title6
ws2.cell(row=1, column=Column_counter+7).value=Sub_title7
ws2.cell(row=1, column=Column_counter+8).value=Sub_title8
ws2.cell(row=1, column=Column_counter+9).value=Sub_title9

Column_Address=utils.get_column_letter(Column_counter+1)  #列番号をアルファベットに変更
ws2.column_dimensions[Column_Address].width=12                #列幅を広げる

wb.save(filepath)

# %%
1028
#データ取得の開始日付入力 print("開始日は2018/12/26以降とする")

Nen_start=input("開始年4桁") 
Tuki_start=input("開始月2桁") 
Niti_start=input("開始日2桁")

print(type(Nen_start), Nen_start, type(Tuki_start), Tuki_start, type(Niti_start), Niti_start)

dt_start_input=datetime.date(int(Nen_start), int(Tuki_start), int(Niti_start))

print("開始日", type(dt_start_input),dt_start_input)

#データ取得の終了日付入力 
Nen_end=input("終了年4桁") 
Tuki_end=input("終了月2桁") 
Niti_end=input("終了日2桁")


dt_end_input=datetime.date(int(Nen_end), int(Tuki_end), int(Niti_end))

print("終了日", type(dt_end_input), dt_end_input)

if dt_end_input<dt_start_input: 
    print("Eroor: 開始日が終了日より後")



# %%
#Sheet4のデータから指定された区間のデータを抜き取りSheet2に張り付ける

from datetime import datetime, date, timedelta

#日付はPython型になっているので変換は不要

#def Change_time(Date11):
#    Date22= datetime.strptime(Date11, '%Y-%m-%d')
#    Date33= date(Date22.year, Date22.month, Date22.day)
#    return Date33
#
k=0
read_row=2
read_column=1
write_row=2

date1=ws4.cell(row=read_row, column=read_column).value #Sheet4の最初の日付
date2=datetime.date(date1)# 時刻を除去し、日付のみとする
print(date1, type(date1), date2, type(date2))

while date2<dt_end_input:
    k=k+1
    if k>3000:
        break
    
    if date2==None:
        print("None")
        break



    if date2>=dt_start_input and date2<=dt_end_input: #読み取った日付が指定の範囲内ならばSheet4のデータをSheet2に転記する
        ws2.cell(row=write_row, column=1).value=date2
        ws2.cell(row=write_row, column=2).value=ws4.cell(row=read_row, column=2).value
        ws2.cell(row=write_row, column=3).value=ws4.cell(row=read_row, column=3).value
        ws2.cell(row=write_row, column=4).value=ws4.cell(row=read_row, column=4).value
        ws2.cell(row=write_row, column=5).value=ws4.cell(row=read_row, column=5).value
        ws2.cell(row=write_row, column=6).value=ws4.cell(row=read_row, column=6).value
        ws2.cell(row=write_row, column=7).value=ws4.cell(row=read_row, column=7).value
        ws2.cell(row=write_row, column=8).value=ws4.cell(row=read_row, column=8).value
        ws2.cell(row=write_row, column=9).value=ws4.cell(row=read_row, column=9).value  
        write_row=write_row+1
        read_row=read_row+1

    else:                         #ifとelseは同じインデントとなる
        read_row=read_row+1
    
    date1=ws4.cell(row=read_row, column=1).value
    print("date1=", date1)
    if date1!="None":
        date2=datetime.date(date1)
        print("date2=", date2)
    else:
        break

print("k:", k)
print("dt_start_input:", dt_start_input)

wb.save(filepath)
#the run results have been confirmd in 03/01/2023

# %%
#グラフにする項目の選択（2項目のみ）

Item_select=input("体重と血糖値なら 1, 体重と血圧なら　2, 血糖値と血圧なら　3, 血糖値と中程度運動なら　4, 中程度運動と運動消費エネルギーなら 5を入力")
print(type(Item_select), Item_select)

if Item_select=="1":
    First_item="体重"
    Second_item="血糖値"
elif Item_select=="2":
    First_item="体重"
    Second_item="血圧"
elif Item_select=="3":
    First_item="血糖値"
    Second_item="血圧"
elif Item_select=="4":
    First_item="血糖値"
    Second_item="中程度運動"
elif Item_select=="5":
    First_item="中程度運動"
    Second_item="運動消費エネルギー"    
else:
    print("Error")
    
print(First_item, Second_item)    

# %%
#2軸のgraphを作成する（折れ線グラフのみ）

from openpyxl.chart import (
    LineChart,
    Reference,
    Series,
)
from openpyxl.chart.axis import DateAxis, GraphicalProperties

#最初グラフオブジェクト（graph_obj1）のｙ軸（左側）

graph_obj1 = LineChart()            #最初のグラフ
graph_obj1.title="健康管理"          #全体のタイトル
graph_obj1.x_axis.title =Sub_title1 #x軸のタイトルは「日付」

graph_obj2 = LineChart()

graph_obj1.y_axis.title =First_item
graph_obj2.y_axis.title =Second_item
    
if Item_select=="1":
    v1 = Reference(ws2, min_col=2, min_row=1, max_col=2, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h1 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択
    v2 = Reference(ws2, min_col=3, min_row=1, max_col=3, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h2 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択
   
    
if Item_select=="2":
    v1 = Reference(ws2, min_col=2, min_row=1, max_col=2, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h1 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択
    v2 = Reference(ws2, min_col=4, min_row=1, max_col=6, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h2 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択 
    

if Item_select=="3":
    v1 = Reference(ws2, min_col=3, min_row=1, max_col=3, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h1 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択
    v2 = Reference(ws2, min_col=4, min_row=1, max_col=6, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h2 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択

    
if Item_select=="4":
    v1 = Reference(ws2, min_col=3, min_row=1, max_col=3, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h1 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択
    v2 = Reference(ws2, min_col=7, min_row=1, max_col=7, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h2 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択

    
if Item_select=="5":
    v1 = Reference(ws2, min_col=7, min_row=1, max_col=7, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h1 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択
    v2 = Reference(ws2, min_col=8, min_row=1, max_col=8, max_row=ws2.max_row)  # Y第1軸の項目含むデータ選択
    h2 = Reference(ws2, min_col=1, min_row=2, max_col=1, max_row=ws2.max_row)  # X軸の項目含まないデータ選択
    
    
graph_obj1.add_data(v1, titles_from_data=True)
graph_obj1.set_categories(h1)  # X軸のグラフ追加
graph_obj2.add_data(v2, titles_from_data=True)
graph_obj2.set_categories(h2)  # X軸のグラフ追加     


# Y第2軸形成

graph_obj2.y_axis.axId = 200  # Y第2軸の形成に関わっているようで実行しないとY第2軸が形成されない
graph_obj2.y_axis.crosses = 'max'  # Y第2軸を右に移動する命令らしい

graph_obj1.width = 25  # グラフのサイズ
graph_obj1.height = 15
graph_obj2.y_axis.majorGridlines = None  # Y第2軸のグリット線を表示しない

# ---- 軸まわりの体裁を強制（数値ラベル表示・重なり回避・日付は下） ----
# ==== 軸表示を強制 & レンジ設定（合成前に実行） ====
from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.chart.shapes import GraphicalProperties

# 角丸・スタイル依存を切る（軸ラベルが消える事象の回避）
for ch in (graph_obj1, graph_obj2):
    ch.roundedCorners = False
    ch.style = None
    ch.graphicalProperties = GraphicalProperties()
    ch.graphicalProperties.line.solidFill = "000000"  # 外枠

# --- 横軸（日付は下／数値表示） ---
graph_obj1.x_axis = DateAxis()                 # 日付軸を明示
graph_obj1.x_axis.delete = False
graph_obj1.x_axis.title = Sub_title1           # 例: "日付"
graph_obj1.x_axis.tickLblPos = "low"           # ラベルを下に
graph_obj1.x_axis.majorGridlines = ChartLines()

# 表示形式（見やすさ優先で「年月」に）
graph_obj1.x_axis.number_format = "yyyy-mm"
graph_obj1.x_axis.majorTimeUnit = "months"
graph_obj1.x_axis.majorUnit = 1               # 1か月ごと
# ※日付（yyyy-mm-dd）を出したい場合は↓に変更:
# graph_obj1.x_axis.number_format = "yyyy-mm-dd"
# graph_obj1.x_axis.majorTimeUnit = "days"; graph_obj1.x_axis.majorUnit = 7  # 7日ごと 等

# --- 左Y軸（体重） ---
graph_obj1.y_axis.delete = False
graph_obj1.y_axis.title = First_item           # 例: "体重"
graph_obj1.y_axis.number_format = "0.0"       # 小数1桁
graph_obj1.y_axis.tickLblPos = "nextTo"
graph_obj1.y_axis.majorGridlines = ChartLines()
graph_obj1.y_axis.crosses = "min"             # 横軸を下固定

# ★ 体重レンジを 50–70 kg に固定
graph_obj1.y_axis.scaling.min = 65
graph_obj1.y_axis.scaling.max = 75
graph_obj1.y_axis.majorUnit = 2               # 2kg刻み（好みで 1/2/5 など）

# タイトルが数値と重ならないように
try:
    graph_obj1.x_axis.title.overlay = False
    graph_obj1.y_axis.title.overlay = False
except Exception:
    pass

# --- 右Y軸（血糖値） ---
graph_obj2.y_axis.delete = False
graph_obj2.y_axis.title = Second_item          # 例: "血糖値"
graph_obj2.y_axis.number_format = "0"          # 整数
graph_obj2.y_axis.tickLblPos = "nextTo"
graph_obj2.y_axis.majorGridlines = None        # 右軸グリッドは消す
graph_obj2.y_axis.axId = 200                   # 2軸化に必須
graph_obj2.y_axis.crosses = "max"              # 右側へ
# 必要なら血糖値側のレンジも固定（例：80–200）
# graph_obj2.y_axis.scaling.min = 80
# graph_obj2.y_axis.scaling.max = 200
# graph_obj2.y_axis.majorUnit = 20

try:
    graph_obj2.y_axis.title.overlay = False
except Exception:
    pass

# 折れ線（スムージング無し）
for s in graph_obj1.series:
    s.smooth = False
for s in graph_obj2.series:
    s.smooth = False


# 別々に作成したY第1軸とY第2軸を1つにする
graph_obj1 += graph_obj2  # 支払額のY第1軸と単価のY第2軸のデータを合成する

ws5.add_chart(graph_obj1, "B2")



wb.save(filepath)

import sys
sys.exit

# %%


# %%



